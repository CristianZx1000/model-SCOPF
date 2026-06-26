from IPython import get_ipython

# Limpiar la consola
ipython = get_ipython()
ipython.run_line_magic('clear', '')  # limpia la consola
ipython.run_line_magic('reset', '-f')  # elimina todas las variables

import warnings

warnings.filterwarnings("ignore", category = DeprecationWarning)

import sys, os
from gurobipy import *
#from scipy.sparse import csr_matrix as sparse
import numpy as np
from numpy import pi#, array, ones, zeros, arange, ix_, r_, flatnonzero as finddiag, dot as mult
#from numpy.linalg import solve, inv
# import pandas as pd
import time
# import matplotlib.pyplot as plt
# import matplotlib.gridspec as gridspec
import pickle
from scipy.stats import truncnorm

import case_SM_PA_V4 as mpc # sep_SM_PA_v5.pdf
import mod_pa_V2 as mpa

# === Configuración caché ===
SCENARIOS_CACHE = "scenarios_data.pkl"
usar_cache = True  # False: para leer desde PF y actualizar caché

if not usar_cache:
    # Enlazar python con PowerFactory # llamar al sistema operativo
    os.environ["PATH"] = r"D:\Program Files\DIgSILENT\PowerFactory 2021 SP2" + os.environ["PATH"]

    # Importar la aplicación # tener acceso desde la ruta donde esta powerfactory
    sys.path.append(r"D:\Program Files\DIgSILENT\PowerFactory 2021 SP2\Python\3.9")

    # Importar la aplicación # importar el módulo que da acceso a la aplicación
    import powerfactory as pf

    app = pf.GetApplication()
    
    # Activar el proyecto
    user = app.GetCurrentUser()                # abre el usuario
    project = app.ActivateProject('BD SM Punta Arenas 2023 ECF y EDAC wind') # abre el archivo pfd
    #project = app.ActivateProject('BD SM Punta Arenas 2023 ECF y EDAC V2')
    prj = app.GetActiveProject()              # activar el proyecto
# 1: Obtener los datos de la red
if usar_cache and os.path.exists(SCENARIOS_CACHE):
    with open(SCENARIOS_CACHE, "rb") as f:
        scenarios_data = pickle.load(f)
else:
    scenarios_data = mpc.get_scenarios_data(app, prj)
    with open(SCENARIOS_CACHE, "wb") as f:
        pickle.dump(scenarios_data, f)

# === Parámetros de costos ===
u_term = 0.7 #1
u_wtg  = 0.1
cdn    = 1 # 0.65

scenarios_data = mpc.compute_costs(scenarios_data, cdn=cdn, u_term=u_term, u_wtg=u_wtg, u_su=0.2)

gen_agc = scenarios_data["sistema"]["gen_agc_info"]

# 2: Construir red
network = mpc.build_network_matrices(scenarios_data, baseMVA=100.0)

dicc_gen_agc = scenarios_data["sistema"]["dicc_gen_agc"]
dicc_gen = {name: i for i, name in enumerate(dicc_gen_agc.keys())}

ng_g = len(dicc_gen_agc)

# True: opt. pre + post, False: opt. solo pre
opt_post = True

# mod. pmax WTGs, 3.45 --> 4 MW 
mod_pmax = False

#⚡⚡⚡⚡⚡⚡⚡⚡⚡⚡⚡⚡⚡⚡⚡⚡⚡⚡⚡⚡⚡⚡⚡⚡⚡⚡⚡⚡
# === Modelo de la incertidumbre ===

p_fore = 3.4 # 3.45 # pronóstico # media: 2, punta: 3.4
if mod_pmax:
    p_fore = 3.8

alfa = 0.1          # factor de la desviación estándar

sigma = alfa * p_fore  # desviación estándar
zeta = 3 * sigma       # límite de desviación
p_VUL = p_fore - zeta  # límite de la parte despachable

# Valor esperado, False: múltiples escenarios de incertidumbre
exp_val = False

if exp_val:
    epsilon_list = np.array([0]) * zeta
else:
    # Parámetros escenarios de epsilon
    rng = np.random.default_rng(42)     # semilla
    # Normal truncada en [-3, 3]
    a, b = -3, 3
    u = truncnorm.rvs(a, b, loc=0, scale=1, size=10, random_state=rng)
    # Normalizar a [-1,1]
    numeros = u / 3 # u_norm

    print(len(numeros))
    print("Media:", round(np.mean(numeros),5), "Std (poblacional):", round(np.std(numeros),5), 
        "Std (muestral):", round(np.std(numeros, ddof=1),5))
    indice_fore = np.argmin(np.abs(numeros))
    valor_fore = numeros[indice_fore]
    print("Valor cercano al esperado", indice_fore, valor_fore)
    idx_min = np.argmin(np.abs(numeros + 1))
    val_min = numeros[idx_min]
    print("Valor mínimo", idx_min, val_min)
    idx_max = np.argmin(np.abs(numeros - 1))
    val_max = numeros[idx_max]
    print("Valor máximo", idx_max, val_max)

    epsilon_list = numeros * zeta

n_w = len(epsilon_list)

eta_list = zeta + epsilon_list     # parte estocástica

t0=time.time()

# ==== Modelo ====
m = Model('DCOPF_3b')
m.setParam('OutputFlag', False)

# Reservas globales
r_up_g = m.addMVar(ng_g, vtype=GRB.CONTINUOUS, lb=0, name='r_up_g')
r_dn_g = m.addMVar(ng_g, vtype=GRB.CONTINUOUS, lb=0, name='r_dn_g')

#########################################################################################################

Cop = LinExpr()
Cto_up_g = np.zeros(ng_g)
Cto_dn_g = np.zeros(ng_g)

casos = list(range(1, 7))
#casos = [4]             # solo un caso
casos_ejecutar = [u for u in casos if u != 3] # Nota: caso 3 se ignora, es igual al caso 2, también se omite en case_SM_PA_V4.py
n_casos = len(casos_ejecutar)
peso_c = 1 / n_casos if n_casos > 1 else 1.0

vars_list = []

# 3: Procesar cada caso
for u in casos_ejecutar:
    # Seleccionar el escenario
    nombre_escenario = f"CASO {u}"

    # Calcular X equivalente con tap
    X_eq = mpc.calculate_reduced_X_trafo_7(scenarios_data, nombre_escenario)
    
    # Calcular matrices
    matrices = mpc.compute_sensitivity_matrices(network, X_eq)
    
    # Preparar datos
    data = mpc.prepare_case_data(scenarios_data, network, nombre_escenario, matrices)

    Sb = data['Sb']
    SL = data['SL']       # slack bus

    ng = data['ng']       # número de gen.
    nb = data['nb']       # número de barras
    nl = data['nl']       # número de líneas

    Pmax = data['Pmax']
    Pmin = data['Pmin']

    # Solo 3 unidades son eólicas en todos los casos, son los últimos tres generadores
    idx_erv = range(ng-3,ng)# [ng-3, ng-2, ng-1]

    if mod_pmax:
        for kr in idx_erv:
            Pmax[kr] *= 4/3.45 # 4
    
    n_erv = len(idx_erv) # 3 gen. eólicos
    eta_vector = np.zeros(ng) # inicialización de eta: parte estocástica

    # Ctos. generadores
    a_g = data['a_g']
    b_g = data['b_g']
    #c_su = data['c_su']
 
    # Nombre de generadores e índice de líneas
    g_names = data['g_names']
    branch_f = data['branch_f']
    branch_t = data['branch_t']

    # Límite reservas
    RUp = RDn = Pmax - Pmin         # límites mín./máx.
    for idx in idx_erv:
        RUp[idx] = RDn[idx] = p_VUL - Pmin[idx]

    # Cto. reserva
    Cto_up_g_caso = data["Cto_up_g"]
    Cto_dn_g_caso = data["Cto_dn_g"]

    idx_activos = [dicc_gen[name] for name in g_names]
    # Guardar costos en vectores globales
    for i_local, i_global in enumerate(idx_activos):
        Cto_up_g[i_global] = Cto_up_g_caso[i_local]
        Cto_dn_g[i_global] = Cto_dn_g_caso[i_local]

    Cg = data['Cg']             # matriz de conexiones

    FM = data['FM']             # F^M
    A = data['A']               # matriz de incidencia
    A_bar = data['A_bar']       # matriz de incidencia no orientada
    BfR = data['BfR']           # yA con pérdidas
    g = data['g']               # conductancia
    b = data['b']               # suceptancia

    idx_con_perdidas = [j for j in range(nl) if g[j] > 1e-9]
    idx_sin_perdidas = [j for j in range(nl) if g[j] <= 1e-9]

    # Carga pre
    Load_bus_pre = data['Load_bus_pre']

    alm_2 = data['alm_2']
    alm_4 = data['alm_4']
    alm_11 = data['alm_11']
    alm_enap = data['alm_enap'] 

    # Generadores que participan en el CSF
    vf = data['vf']     # parámetro binario
    
    # Número de tramos para la linealización de las pérdidas
    L = 6

    k_coef = np.zeros((nl,L))
    for l in range(1,L+1):
        k_coef[:,l-1] = (2*l-1) * FM/(Sb*L)

    # Cond. de complementariedad    
    compl = True

    # Cond. de adyacencia
    ady = True

    # ENS
    Cto_ens = 10.04e3
    Pmax_ens = Load_bus_pre
    
    contingencias = [('gen', i+1) for i in range(ng)]
    contingencias += [('load', 2), ('load', 4), ('load', 11), ('load', 15)]
    
    K = len(contingencias)

    #####################################################################################################################################
    # Definición de variables

    # Precontigencia
    
    p_pre = m.addMVar((ng, n_w), vtype=GRB.CONTINUOUS, lb=0, name=f'Pg_pre_caso{u}') #lb lim inf p>=0
    d_pre = m.addMVar((nb, n_w), vtype=GRB.CONTINUOUS, lb=-GRB.INFINITY, ub=GRB.INFINITY, name=f'd_pre_caso{u}')
    p_ens_pre = m.addMVar((nb, n_w), vtype=GRB.CONTINUOUS, lb=0, name=f'p_ens_pre_caso{u}')

    f_pre = m.addMVar((nl, n_w), vtype=GRB.CONTINUOUS, lb=-GRB.INFINITY, ub=GRB.INFINITY, name=f'f_pre_caso{u}')
    fp_pre = m.addMVar((nl, n_w), vtype=GRB.CONTINUOUS, lb=0, name=f'fp_pre_caso{u}')
    fn_pre = m.addMVar((nl, n_w), vtype=GRB.CONTINUOUS, lb=0, name=f'fn_pre_caso{u}')
    ploss_pre = m.addMVar((nl, n_w), vtype=GRB.CONTINUOUS, lb=0, name=f'Ploss_pre_caso{u}')
    df_pre = m.addMVar((nl, L, n_w), vtype=GRB.CONTINUOUS, lb=0, name=f'df_pre_caso{u}')
    n_lf_pre = m.addMVar((nl, n_w), vtype=GRB.BINARY, name=f'n_lf_pre_caso{u}') # Var. bin. cond. de complementariedad
    n_df_pre = m.addMVar((nl, L, n_w), vtype=GRB.BINARY, name=f'n_df_pre_caso{u}') # Var. bin. cond. de adyacencia

    if len(idx_sin_perdidas) > 0:
        df_pre[idx_sin_perdidas, :, :].ub = 0
        n_lf_pre[idx_sin_perdidas, :].ub = 0
        n_df_pre[idx_sin_perdidas, :, :].ub = 0
        fp_pre[idx_sin_perdidas, :].ub = 0
        fn_pre[idx_sin_perdidas, :].ub = 0

    # Postcontingencia

    p_post = m.addMVar((ng, K, n_w), vtype=GRB.CONTINUOUS, lb=0, name=f'Pg_post_caso{u}')
    d_post = m.addMVar((nb, K, n_w), vtype=GRB.CONTINUOUS, lb=-GRB.INFINITY, ub=GRB.INFINITY, name=f'd_post_caso{u}')
    p_ens_post = m.addMVar((nb, K, n_w), vtype=GRB.CONTINUOUS, lb=0, name=f'p_ens_post_caso{u}')
    
    f_post = m.addMVar((nl, K, n_w), vtype=GRB.CONTINUOUS, lb=-GRB.INFINITY, ub=GRB.INFINITY, name=f'f_post_caso{u}')
    fp_post = m.addMVar((nl, K, n_w), vtype=GRB.CONTINUOUS, lb=0, name=f'fp_post_caso{u}')
    fn_post = m.addMVar((nl, K, n_w), vtype=GRB.CONTINUOUS, lb=0, name=f'fn_post_caso{u}')
    ploss_post = m.addMVar((nl, K, n_w), vtype=GRB.CONTINUOUS, lb=0, name=f'Ploss_post_caso{u}')
    df_post = m.addMVar((nl, L, K, n_w), vtype=GRB.CONTINUOUS, lb=0, name=f'df_post_caso{u}')
    n_lf_post = m.addMVar((nl, K, n_w), vtype=GRB.BINARY, name=f'n_lf_post_caso{u}')
    n_df_post = m.addMVar((nl, L, K, n_w), vtype=GRB.BINARY, name=f'n_df_post_caso{u}')

    if len(idx_sin_perdidas) > 0:
        df_post[idx_sin_perdidas, :, :, :].ub = 0
        n_lf_post[idx_sin_perdidas, :, :].ub = 0
        n_df_post[idx_sin_perdidas, :, :, :].ub = 0
        fp_post[idx_sin_perdidas, :, :].ub = 0
        fn_post[idx_sin_perdidas, :, :].ub = 0

    # Variable de encendido de gen.
    u_i = m.addMVar(ng, vtype=GRB.BINARY, name=f'u_i_caso{u}')
    m.addConstrs((u_i[h] == 1 for h in idx_erv), name=f'u_i_caso{u}_ones')

    # Reserva caso u
    r_up = m.addMVar(ng, vtype=GRB.CONTINUOUS, lb=0, name=f'r_up_caso{u}')
    r_dn = m.addMVar(ng, vtype=GRB.CONTINUOUS, lb=0, name=f'r_dn_caso{u}')

    # Acoplamiento con reservas globales
    for i_local, i_global in enumerate(idx_activos):
        m.addConstr(r_up[i_local] == r_up_g[i_global])
        m.addConstr(r_dn[i_local] == r_dn_g[i_global])
    
    #####################################################################################################################################

    # ==== Función objetivo ====
    
    Cop_pre_total = LinExpr()
    Cop_post_total = LinExpr()

    prob_w = 1 / n_w  # equiprobables

    for w in range(n_w):
        # Precontingencia para escenario w
        Cop_pre_w = a_g @ p_pre[:, w] + b_g @ u_i + p_ens_pre[:, w].sum() * Cto_ens
        #Cop_pre_w = a_g @ p_pre[:, w] + b_g @ u_i + c_su @ u_i + p_ens_pre[:, w].sum() * Cto_ens
        Cop_pre_total += prob_w * Cop_pre_w
    
        # Postcontingencia para escenario w (promedio sobre todas las contingencias)
        for k in range(K):
            if opt_post:
                Cop_post_w_k = a_g @ p_post[:, k, w] + b_g @ u_i + p_ens_post[:, k, w].sum() * Cto_ens
            else:
                Cop_post_w_k = p_ens_post[:, k, w].sum() * Cto_ens
            Cop_post_total += prob_w * Cop_post_w_k / K

    # Contribución del caso u a la FO total
    Cop += peso_c * (Cop_pre_total + Cop_post_total)

    vars_case = {
        'p_pre': p_pre,
        'f_pre': f_pre,
        'FM': FM,
        'a_g': a_g,
        #'c_su': c_su,
        'u_i': u_i,
        'ploss_pre': ploss_pre,
        'p_ens_pre': p_ens_pre,
        'p_post': p_post,
        'f_post': f_post,
        'ploss_post': ploss_post,
        'p_ens_post': p_ens_post,
        'Cop_pre': Cop_pre_total,
        'Cop_post': Cop_post_total,
        'gen_names': g_names,
        'branch_from': branch_f,
        'branch_to': branch_t,
        'contingencias': contingencias,
        'Load_bus_pre': Load_bus_pre,
        'Load_bus_post': {}
    }

    # ==== Subjet to: ====
    
    for w in range(n_w):
        eta = eta_list[w]
        for idx_w in idx_erv:
            eta_vector[idx_w] = eta

        p_pre_w = p_pre[:, w]
        d_pre_w = d_pre[:, w]
        p_ens_pre_w = p_ens_pre[:, w]
        f_pre_w = f_pre[:, w]
        fp_pre_w = fp_pre[:, w]
        fn_pre_w = fn_pre[:, w]
        ploss_pre_w = ploss_pre[:, w]
        df_pre_w = df_pre[:, :, w]
        n_lf_pre_w = n_lf_pre[:, w]
        n_df_pre_w = n_df_pre[:, :, w]

        #####################################################################################################################################
        # ==== Precontingencia ====

        # Barra SL
        m.addConstr(d_pre_w[SL] == 0, f'SL_pre_caso{u}_w{w}')

        # Balance (LCK)
        m.addConstr(Cg @ p_pre_w + Cg @ eta_vector + p_ens_pre_w - Load_bus_pre - 0.5 * A_bar.T @ ploss_pre_w == A.T @ f_pre_w, name=f'LCK_pre_caso{u}_w{w}') #Cg*Pg+P_ens-D-0.5*A^T*P_loss=A^T*f

        # Límite ángulos
        m.addConstr(-A @ d_pre_w >= -pi/2, name=f'dM_pre_caso{u}_w{w}')
        m.addConstr(A @ d_pre_w >= -pi/2, name=f'dm_pre_caso{u}_w{w}')

        # P_max y P_min
        for t in range(ng-3):
            m.addConstr(p_pre_w[t] >= Pmin[t] * u_i[t], name=f'P_min_pre{t}_caso{u}_w{w}')
            m.addConstr(-p_pre_w[t] >= -Pmax[t] * u_i[t], name=f'P_max_pre{t}_caso{u}_w{w}')
        for idx_e in idx_erv:
            m.addConstr(p_pre_w[idx_e] + eta_vector[idx_e] >= Pmin[idx_e], name=f'P_min_pre{idx_e}_caso{u}_w{w}')
            m.addConstr(-p_pre_w[idx_e] - eta_vector[idx_e] >= -Pmax[idx_e], name=f'P_max_pre{idx_e}_caso{u}_w{w}')
        
        # Límite VUL
            m.addConstr(-p_pre_w[idx_e] - r_up[idx_e] >= -p_VUL, name = f'P_base_VUL{idx_e}_caso{u}_{w}')

        # Reserva
        m.addConstr(-r_up >= -RUp * vf, name=f'RUp_caso{u}_w{w}')
        m.addConstr(-r_dn >= -RDn * vf, name=f'RDn_caso{u}_w{w}')

        # ENS
        m.addConstr(-p_ens_pre_w >= -Pmax_ens, name=f'P_max_ens_pre_caso{u}_w{w}')

        #####################################################################################################################################
        # Límites y pérdidas líneas pre

        # --- 1. Cálculo de pérdidas ---
        for j in idx_con_perdidas:
            sum_kdf_pre = df_pre_w[j, :] @ k_coef[j, :]
            m.addConstr(ploss_pre_w[j] == (g[j]/b[j]**2) * sum_kdf_pre, name=f'f_lin_{j}_pre_caso{u}_w{w}')

        for j in idx_sin_perdidas:
            # Solo fijar las pérdidas a 0
            m.addConstr(ploss_pre_w[j] == 0, name=f'ploss_zero_{j}_pre_caso{u}_w{w}')

        # --- 2. Límites de los tramos (solo ramas con pérdidas) ---
        if not ady:
            for l in range(0, L):
                m.addConstr(-df_pre_w[idx_con_perdidas, l] >= -FM[idx_con_perdidas]/L, name=f'df_max_pre_{l}_caso{u}_w{w}') 

        # LVK y límites
        m.addConstr(f_pre_w == -Sb * BfR @ d_pre_w, name = f'LVK_pre_caso{u}_w{w}')
        m.addConstr(-f_pre_w >= -FM, name = f'fmax_pre_caso{u}_w{w}')
        m.addConstr(f_pre_w >= -FM, name = f'fmin_pre_caso{u}_w{w}')

        # --- 3. Condición de Complementariedad (SOLO ramas con pérdidas) ---
        if not compl:
            for j in idx_con_perdidas:
                m.addConstr(-fp_pre_w[j] >= -FM[j], name = f'fp_max_pre_caso{u}_w{w}_{j}')
                m.addConstr(-fn_pre_w[j] >= -FM[j], name = f'fn_max_pre_caso{u}_w{w}_{j}')
        else:
            # Solo aplicar las binarias n_lf a líneas con pérdidas
            # En las líneas sin pérdidas, restricción sin variables binarias
            for j in idx_con_perdidas:
                m.addConstr(-fp_pre_w[j] >= -FM[j] * n_lf_pre_w[j], name = f'fp_max_pre_c_{j}_caso{u}_w{w}')
                m.addConstr(-fn_pre_w[j] >= -FM[j] * (1 - n_lf_pre_w[j]), name = f'fn_max_pre_c_{j}_caso{u}_w{w}')

        # --- 4. Relaciones de flujo (SOLO ramas con pérdidas) ---
        for j in idx_con_perdidas:
            m.addConstr(df_pre_w[j, :].sum() == fp_pre_w[j] + fn_pre_w[j], name = f'f0_pre_{j}_caso{u}_w{w}')
            m.addConstr(f_pre_w[j] == fp_pre_w[j] - fn_pre_w[j], name = f'f1_pre_{j}_caso{u}_w{w}')
            m.addConstr(-f_pre_w[j] - 0.5*ploss_pre_w[j] >= -FM[j], name = f'fmax_loss_pre_{j}_caso{u}_w{w}')
            m.addConstr(f_pre_w[j] - 0.5*ploss_pre_w[j] >= -FM[j], name = f'fmin_loss_pre_{j}_caso{u}_w{w}')

        # --- 5. Adyacencia (SOLO ramas con pérdidas) ---
        if ady:
            for j in idx_con_perdidas:
                for l in range(L):  
                    if l == 0:
                        m.addConstr(-df_pre_w[j, l] >= -FM[j]/L, name=f'df_pre_{j}_{l}_max_caso{u}_w{w}')
                        m.addConstr(df_pre_w[j, l] >= n_df_pre_w[j, l] * FM[j]/L, name=f'df_pre_{j}_{l}_min_caso{u}_w{w}')
                    elif l == L-1:
                        m.addConstr(-df_pre_w[j, l] >= -n_df_pre_w[j, l-1] * FM[j]/L, name=f'df_pre_{j}_{l}_max_caso{u}_w{w}')
                    else:        
                        m.addConstr(-df_pre_w[j, l] >= -n_df_pre_w[j, l-1] * FM[j]/L, name=f'df_pre_{j}_{l}_max_caso{u}_w{w}')
                        m.addConstr(df_pre_w[j, l] >= n_df_pre_w[j, l] * FM[j]/L, name=f'df_pre_{j}_{l}_min_caso{u}_w{w}')

        #####################################################################################################################################
        # ==== Postcontingencia ====
        
        for k_idx,(tipo,index) in enumerate(contingencias):
            p_post_k     = p_post[:,k_idx, w]
            d_post_k     = d_post[:,k_idx, w]
            p_ens_post_k = p_ens_post[:,k_idx, w]
            f_post_k     = f_post[:,k_idx, w]
            fp_post_k    = fp_post[:,k_idx, w]
            fn_post_k    = fn_post[:,k_idx, w]
            ploss_post_k = ploss_post[:,k_idx, w]
            df_post_k    = df_post[:,:,k_idx, w]
            n_lf_post_k  = n_lf_post[:,k_idx, w]
            n_df_post_k  = n_df_post[:,:,k_idx, w]
            eta_vector_post = eta_vector.copy()

            # N-1 cargas
            if tipo == 'load':
                Load_bus_post_k = Load_bus_pre.copy()
                if index == 2:    
                    Load_bus_post_k[5] -= alm_2
                elif index == 4:    
                    Load_bus_post_k[1] -= alm_4
                elif index == 11:    
                    Load_bus_post_k[2] -= alm_11
                else:    
                    Load_bus_post_k[7] -= alm_enap

            # N-1 gen.
            elif tipo == 'gen':
                # Anular el componente estocástico para el generador fuera de servicio
                Load_bus_post_k = Load_bus_pre.copy()
                eta_vector_post[index-1] = 0
            
            nombre_k = f"Cont{k_idx+1}_{tipo}{index}"
            vars_case['Load_bus_post'][nombre_k] = Load_bus_post_k

            # Barra SL
            m.addConstr(d_post_k[SL] == 0, f'SL_post[{k_idx}]_caso{u}_w{w}')

            # Balance (LCK)
            m.addConstr(Cg @ p_post_k + Cg @ eta_vector_post + p_ens_post_k - Load_bus_post_k - 0.5 * A_bar.T @ ploss_post_k == A.T @ f_post_k, name = f'LCK_post[{k_idx}]_caso{u}_w{w}')

            #####################################################################################################################################
            # Límites y pérdidas líneas post
            
            # --- 1. Cálculo de pérdidas ---
            for j in idx_con_perdidas:
                sum_kdf_post = df_post_k[j, :] @ k_coef[j, :]
                m.addConstr(ploss_post_k[j] == (g[j]/b[j]**2) * sum_kdf_post, name=f'f_lin_{j}_post[{k_idx}]_caso{u}_w{w}')

            for j in idx_sin_perdidas:
                # Se mantienen fijas en cero
                m.addConstr(ploss_post_k[j] == 0, name=f'ploss_zero_{j}_post[{k_idx}]_caso{u}_w{w}')

            # --- 2. Límites de los tramos (solo ramas con pérdidas) ---
            if not ady:
                for l in range(0, L):
                    m.addConstr(-df_post_k[idx_con_perdidas, l] >= -FM[idx_con_perdidas]/L, name=f'df_max_post_{l}[{k_idx}]_caso{u}_w{w}') 

            # LVK y límites
            m.addConstr(f_post_k == -Sb * BfR @ d_post_k, name = f'LVK_post[{k_idx}]_caso{u}_w{w}')
            m.addConstr(-f_post_k >= -FM, name = f'fmax_post[{k_idx}]_caso{u}_w{w}')
            m.addConstr(f_post_k >= -FM, name = f'fmin_post[{k_idx}]_caso{u}_w{w}')

            # --- 3. Condición de Complementariedad (SOLO ramas con pérdidas) ---
            if not compl:
                for j in idx_con_perdidas:
                    m.addConstr(-fp_post_k[j] >= -FM[j], name = f'fp_max_post_{j}[{k_idx}]_caso{u}_w{w}')
                    m.addConstr(-fn_post_k[j] >= -FM[j], name = f'fn_max_post_{j}[{k_idx}]_caso{u}_w{w}')
            else:
                for j in idx_con_perdidas:
                    m.addConstr(-fp_post_k[j] >= -FM[j] * n_lf_post_k[j], name = f'fp_max_post_c_{j}[{k_idx}]_caso{u}_w{w}')
                    m.addConstr(-fn_post_k[j] >= -FM[j] * (1 - n_lf_post_k[j]), name = f'fn_max_post_c_{j}[{k_idx}]_caso{u}_w{w}')
                    
            # --- 4. Relaciones de Flujo (SOLO ramas con pérdidas) ---
            for j in idx_con_perdidas:
                m.addConstr(df_post_k[j, :].sum() == fp_post_k[j] + fn_post_k[j], name = f'f0_post_{j}[{k_idx}]_caso{u}_w{w}')
                m.addConstr(f_post_k[j] == fp_post_k[j] - fn_post_k[j], name = f'f1_post_{j}[{k_idx}]_caso{u}_w{w}')
                m.addConstr(-f_post_k[j] - 0.5*ploss_post_k[j] >= -FM[j], name = f'fmax_loss_post_{j}[{k_idx}]_caso{u}_w{w}')
                m.addConstr(f_post_k[j] - 0.5*ploss_post_k[j] >= -FM[j], name = f'fmin_loss_post_{j}[{k_idx}]_caso{u}_w{w}')

            # --- 5. Adyacencia (SOLO para ramas con pérdidas) ---
            if ady:
                for j in idx_con_perdidas:
                    for l in range(L):  
                        if l == 0:
                            m.addConstr(-df_post_k[j, l] >= -FM[j]/L, name=f'df_post_{j}_{l}_max[{k_idx}]_caso{u}_w{w}')
                            m.addConstr(df_post_k[j, l] >= n_df_post_k[j, l] * FM[j]/L, name=f'df_post_{j}_{l}_min[{k_idx}]_caso{u}_w{w}')
                        elif l == L-1:
                            m.addConstr(-df_post_k[j, l] >= -n_df_post_k[j, l-1] * FM[j]/L, name=f'df_post_{j}_{l}_max[{k_idx}]_caso{u}_w{w}')
                        else:        
                            m.addConstr(-df_post_k[j, l] >= -n_df_post_k[j, l-1] * FM[j]/L, name=f'df_post_{j}_{l}_max[{k_idx}]_caso{u}_w{w}')
                            m.addConstr(df_post_k[j, l] >= n_df_post_k[j, l] * FM[j]/L, name=f'df_post_{j}_{l}_min[{k_idx}]_caso{u}_w{w}')

            #####################################################################################################################################
            # Límite ángulos
            m.addConstr(-A @ d_post_k >= -pi/2, name = f'dM_post[{k_idx}]_caso{u}_w{w}')
            m.addConstr(A @ d_post_k >= -pi/2, name = f'dm_post[{k_idx}]_caso{u}_w{w}')

            # Generador fuera de servicio
            if tipo == 'gen':
                m.addConstr(p_post_k[index-1] == 0, f'Out_service[{k_idx}]_caso{u}_w{w}')
                for h in range(ng):
                    if h != index-1:   # todos excepto el fuera de servicio
                        # P_max y P_min
                        if h not in idx_erv:
                            m.addConstr(p_post_k[h] >= Pmin[h] * u_i[h], name=f'Pmin_post[{k_idx},{h}]_caso{u}_w{w}')
                            m.addConstr(-p_post_k[h] >= -Pmax[h] * u_i[h], name=f'Pmax_post[{k_idx},{h}]_caso{u}_w{w}')
                        if h in idx_erv:
                            m.addConstr(p_post_k[h] + eta_vector_post[h] >= Pmin[h], name=f'Pmin_post[{k_idx},{h}]_caso{u}_w{w}')
                            m.addConstr(-p_post_k[h] - eta_vector_post[h] >= -Pmax[h], name=f'Pmax_post[{k_idx},{h}]_caso{u}_w{w}')
                                               
                        #  Gen. renovables (límite VUL)
                            m.addConstr(-p_post_k[h] >= -p_VUL, name = f'P_base_VUL[{k_idx},{h}]_caso{u}_w{w}')

                        # Reserva
                        m.addConstr(p_pre_w[h] + r_up[h] >= p_post_k[h], name=f'Up[{k_idx},{h}]_caso{u}_w{w}')
                        m.addConstr(-p_pre_w[h] + r_dn[h] >= -p_post_k[h], name=f'Dn[{k_idx},{h}]_caso{u}_w{w}')
            
            else:
                # P_max y P_min
                for t in range(ng-3):
                    m.addConstr(p_post_k[t] >= Pmin[t] * u_i[t], name = f'P_min_post[{k_idx},{t}]_caso{u}_w{w}')
                    m.addConstr(-p_post_k[t] >= -Pmax[t] * u_i[t], name= f'P_max_post[{k_idx},{t}]_caso{u}_w{w}')
                for idx_e in idx_erv:
                    m.addConstr(p_post_k[idx_e] + eta_vector_post[idx_e] >= Pmin[idx_e], name = f'P_min_post[{k_idx},{idx_e}]_caso{u}_w{w}')
                    m.addConstr(-p_post_k[idx_e] - eta_vector_post[idx_e] >= -Pmax[idx_e], name= f'P_max_post[{k_idx},{idx_e}]_caso{u}_w{w}')

                # Gen. renovables (límite VUL)
                    m.addConstr(-p_post_k[idx_e] >= -p_VUL, name = f'P_base_VUL[{k_idx},{idx_e}]_caso{u}_w{w}')

                # Reserva
                m.addConstr(p_pre_w + r_up >= p_post_k, name = f'Up[{k_idx}]_caso{u}_w{w}')
                m.addConstr(-p_pre_w + r_dn >= -p_post_k, name = f'Dn[{k_idx}]_caso{u}_w{w}')

            # ENS
            m.addConstr(-p_ens_post_k >= -Pmax_ens, name=f'P_max_ens_post[{k_idx}]_caso{u}_w{w}')
        
    vars_list.append(vars_case)

C_res = r_up_g @ Cto_up_g + r_dn_g @ Cto_dn_g

Cop += C_res

m.setObjective(Cop, GRB.MINIMIZE)

t1=time.time()
m.optimize()
t2=time.time()

status = m.Status
if status == GRB.Status.OPTIMAL:
    print('Optimal found => status "%d"' % status)
elif status == GRB.Status.INF_OR_UNBD or \
    status == GRB.Status.INFEASIBLE  or \
    status == GRB.Status.UNBOUNDED:
    print('The model cannot be solved because it is infeasible or unbounded => status "%d"' % status)
if not status == GRB.Status.OPTIMAL:
    ejecutar = False
    assert ejecutar, "Ejecución detenida por condición"

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
import numpy as np
import os

nombre_archivo = "resultados_SM_PA_v7_estocastico.xlsx"
# Diccionario para agregar nombres descriptivos a líneas específicas
# Se incluyen ambos sentidos (ida y vuelta) para asegurar el match
descripciones_lineas = {
    (4, 5): "(66 kV)", (5, 4): "(66 kV)",
    (7, 8): "(PE)", (8, 7): "(PE)",
    (8, 9): "(PE)", (9, 8): "(PE)",
    (9, 10): "(PE)", (10, 9): "(PE)",
    (10, 12): "(WTG1-WTG2)", (12, 10): "(WTG1-WTG2)",
    (12, 14): "(WTG2-WTG3)", (14, 12): "(WTG2-WTG3)"
}

# 1. Preparar DataFrame de reservas globales
datos_reservas = []
for name in gen_agc:
    matches = [key for key in dicc_gen.keys() if name in key]
    i = dicc_gen[matches[0]]
    datos_reservas.append({
        "Generador": name,
        "Reserva Up (MW)": round(float(r_up_g.X[i]), 3),
        "Reserva Dn (MW)": round(float(r_dn_g.X[i]), 3)
    })

# Calcular la suma de las últimas 3 filas (correspondientes a los PE)
suma_up_pe = sum(fila["Reserva Up (MW)"] for fila in datos_reservas[-3:])
suma_dn_pe = sum(fila["Reserva Dn (MW)"] for fila in datos_reservas[-3:])

# Añadir la fila con el total al final de la lista
datos_reservas.append({
    "Generador": "Reserva total PE",
    "Reserva Up (MW)": round(suma_up_pe, 3),
    "Reserva Dn (MW)": round(suma_dn_pe, 3)
})

df_reservas = pd.DataFrame(datos_reservas)

# ExcelWriter
with pd.ExcelWriter(nombre_archivo, engine="openpyxl") as writer:
    
    # 2. Iterar sobre cada CASO operativo (u)
    for u_idx, vars_case in enumerate(vars_list):
        #nombre_hoja = f"Caso {u_idx+1}"
        nombre_hoja = f"Caso {casos_ejecutar[u_idx]}"
        
        # Recuperar variables del caso
        p_pre = vars_case['p_pre']
        ploss_pre = vars_case['ploss_pre']
        p_post = vars_case['p_post']
        ploss_post = vars_case['ploss_post']
        
        # Variables adicionales
        p_ens_pre = vars_case.get('p_ens_pre')
        p_ens_post = vars_case.get('p_ens_post')
        f_pre = vars_case.get('f_pre')
        f_post = vars_case.get('f_post')
        branch_from = vars_case.get('branch_from')
        branch_to = vars_case.get('branch_to')
        
        # Convertir costos a float
        try:
            val_pre = vars_case['Cop_pre'].getValue()
            val_post = vars_case['Cop_post'].getValue()
            Cop_pre_val = float(val_pre) if np.ndim(val_pre) == 0 else float(val_pre.item())
            Cop_post_val = float(val_post) if np.ndim(val_post) == 0 else float(val_post.item())
        except:
            Cop_pre_val = float(vars_case['Cop_pre'].getValue())
            Cop_post_val = float(vars_case['Cop_post'].getValue())
        
        Load_bus_pre = vars_case['Load_bus_pre']
        Load_bus_post = vars_case['Load_bus_post']
        
        gen_names = vars_case['gen_names']
        contingencias = vars_case['contingencias']
        
        ng = p_pre.shape[0]
        n_w_caso = p_pre.shape[1]
        K = p_post.shape[1]
        
        # Obtener dimensiones si existen las variables
        nb = p_ens_pre.shape[0] if p_ens_pre is not None else 0
        nl = f_pre.shape[0] if f_pre is not None else 0

        # --- ESCRIBIR ENCABEZADOS Y RESERVAS ---
        fila_actual = 0
        
        # Título del caso
        pd.DataFrame([f"Resultados {nombre_hoja}"]).to_excel(
            writer, sheet_name=nombre_hoja, startrow=fila_actual, 
            startcol=0, index=False, header=False
        )
        fila_actual += 1
        
        # Tabla de costos esperados
        df_costos = pd.DataFrame([
            {"Parámetro": "Costo total (valor esperado) ($/h)", "Valor": round(m.objVal, 2)},
            {"Parámetro": "Costo precontingencia (valor esperado)", "Valor": round(Cop_pre_val, 2)},
            {"Parámetro": "Costo reserva", "Valor": round(float(C_res.getValue()), 2)},
            {"Parámetro": "Costo reserva Up", "Valor": round(float(r_up_g.X @ Cto_up_g), 2)},
            {"Parámetro": "Costo reserva Dn", "Valor": round(float(r_dn_g.X @ Cto_dn_g), 2)},
            {"Parámetro": "Costo postcontingencia (valor esperado)", "Valor": round(Cop_post_val, 2)},
            {"Parámetro": "Pronóstico (p_fore) [MW]", "Valor": round(p_fore, 2)},
            {"Parámetro": "Límite VUL (p_VUL) [MW]", "Valor": round(p_VUL, 2)}
        ])
        df_costos.to_excel(writer, sheet_name=nombre_hoja, startrow=fila_actual, startcol=0, index=False)
        
        # Tabla de reservas
        df_reservas.to_excel(writer, sheet_name=nombre_hoja, startrow=fila_actual, startcol=4, index=False)
        
        ws = writer.sheets[nombre_hoja]
        
        # El encabezado está en fila_actual + 1, por lo que la última fila de datos
        # está exactamente en fila_actual + 1 + len(df_reservas)
        fila_total_excel = fila_actual + 1 + len(df_reservas)
        
        # startcol=4 corresponde a la columna 5 (E). Modificamos las columnas E, F y G (5, 6 y 7)
        for col_idx in range(5, 8):
            ws.cell(row=fila_total_excel, column=col_idx).font = Font(name="Calibri", bold=True)
        
        fila_actual += max(len(df_costos), len(df_reservas)) + 2

        # ====================================================================
        # CLASIFICACIÓN: LÍNEAS VS TRAFOS
        # Todo lo que esté en descripciones_lineas es Línea, el resto Trafo
        # ====================================================================
        lineas_idx = []
        trafos_idx = []
        
        for l in range(nl):
            bus_from = int(branch_from[l])
            bus_to = int(branch_to[l])
            
            es_linea = (bus_from, bus_to) in descripciones_lineas
            
            if es_linea:
                lineas_idx.append(l)
            else:
                trafos_idx.append(l)
        # ====================================================================

        # 3. Iterar sobre cada escenario de incertidumbre (w)
        for w in range(n_w_caso):
            epsilon_val = epsilon_list[w]
            eta_val = eta_list[w]
            
            # Encabezado escenario incertidumbre
            header_w = f"Escenario {w+1} (incertidumbre: {epsilon_val:.3f} MW | eta: {eta_val:.3f} MW)"
            pd.DataFrame([header_w]).to_excel(
                writer, sheet_name=nombre_hoja, startrow=fila_actual, 
                startcol=0, index=False, header=False
            )
            fila_actual += 1
            
            # ========================================
            # TABLA: PRECONTINGENCIA (w)
            # ========================================
            datos_pre = []
            for h in range(ng):
                clean_name = gen_names[h].replace('.ElmSym', '').replace('.ElmGenstat', '')
                datos_pre.append({
                    "Generador": clean_name,
                    "Potencia pre (MW)": round(float(p_pre.X[h, w]), 3)
                })
            
            datos_pre.append({
                "Generador": "Pérdidas totales",
                "Potencia pre (MW)": round(float(ploss_pre.X[:, w].sum()), 3)
            })
            
            # ENS precontingencia
            if p_ens_pre is not None:
                ens_pre_total = sum(p_ens_pre.X[b, w] for b in range(nb))
                datos_pre.append({
                    "Generador": "ENS total",
                    "Potencia pre (MW)": round(float(ens_pre_total), 3)
                })
            
            datos_pre.append({
                "Generador": "Carga total",
                "Potencia pre (MW)": round(float(sum(Load_bus_pre)), 3)
            })
            
            df_pre_w = pd.DataFrame(datos_pre)
            df_pre_w.to_excel(writer, sheet_name=nombre_hoja, startrow=fila_actual, startcol=0, index=False)
            fila_actual += len(df_pre_w) + 2
            
            # ========================================
            # TABLA: POSTCONTINGENCIA (w)
            # ========================================
            pd.DataFrame(["Postcontingencia"]).to_excel(
                writer, sheet_name=nombre_hoja, startrow=fila_actual-1, 
                startcol=0, index=False, header=False
            )
            
            datos_post_dict = {}
            lista_generadores = [gen_names[h].replace('.ElmSym', '').replace('.ElmGenstat', '') for h in range(ng)]
            datos_post_dict["Generador\ contingencia"] = lista_generadores
            
            for c in range(K):
                tipo, idx = contingencias[c]
                #nombre_col = f"Cont{c+1}: {tipo}{idx}"
                if tipo == "gen":
                    nombre_col = f"Cont{c+1}: gen. {idx}"
                else:
                    if idx == 2:
                        nombre_col = f"Cont{c+1}: alm. 2"
                    elif idx == 4:
                        nombre_col = f"Cont{c+1}: alm. 4"
                    else:
                        nombre_col = f"Cont{c+1}: alm. ENAP"
                datos_post_dict[nombre_col] = [round(float(p_post.X[h, c, w]), 3) for h in range(ng)]
            
            df_post_w = pd.DataFrame(datos_post_dict)
            
            # Fila de pérdidas
            fila_ploss = {"Generador\ contingencia": "Pérdidas (MW)"}
            for c in range(K):
                tipo, idx = contingencias[c]
                #nombre_col = f"Cont{c+1}: {tipo}{idx}"
                if tipo == "gen":
                    nombre_col = f"Cont{c+1}: gen. {idx}"
                else:
                    if idx == 2:
                        nombre_col = f"Cont{c+1}: alm. 2"
                    elif idx == 4:
                        nombre_col = f"Cont{c+1}: alm. 4"
                    else:
                        nombre_col = f"Cont{c+1}: alm. ENAP"
                fila_ploss[nombre_col] = round(float(ploss_post.X[:, c, w].sum()), 3)
            
            # Fila ENS total
            if p_ens_post is not None:
                fila_ens = {"Generador\ contingencia": "ENS total (MW)"}
                for c in range(K):
                    tipo, idx = contingencias[c]
                    #nombre_col = f"Cont{c+1}: {tipo}{idx}"
                    if tipo == "gen":
                        nombre_col = f"Cont{c+1}: gen. {idx}"
                    else:
                        if idx == 2:
                            nombre_col = f"Cont{c+1}: alm. 2"
                        elif idx == 4:
                            nombre_col = f"Cont{c+1}: alm. 4"
                        else:
                            nombre_col = f"Cont{c+1}: alm. ENAP"
                    ens_total = p_ens_post.X[:, c, w].sum()
                    fila_ens[nombre_col] = round(float(ens_total), 3)
            
            # Fila de carga
            fila_carga = {"Generador\ contingencia": "Carga total (MW)"}
            for c in range(K):
                tipo, idx = contingencias[c]
                nombre_k = f"Cont{c+1}_{tipo}{idx}"
                Load_bus_k = Load_bus_post[nombre_k]
                #nombre_col = f"Cont{c+1}: {tipo}{idx}"
                if tipo == "gen":
                    nombre_col = f"Cont{c+1}: gen. {idx}"
                else:
                    if idx == 2:
                        nombre_col = f"Cont{c+1}: alm. 2"
                    elif idx == 4:
                        nombre_col = f"Cont{c+1}: alm. 4"
                    else:
                        nombre_col = f"Cont{c+1}: alm. ENAP"
                fila_carga[nombre_col] = round(float(sum(Load_bus_k)), 3)
            
            # Concatenar todas las filas
            if p_ens_post is not None:
                df_post_w = pd.concat([
                    df_post_w, 
                    pd.DataFrame([fila_ploss]), 
                    pd.DataFrame([fila_ens]),
                    pd.DataFrame([fila_carga])
                ], ignore_index=True)
            else:
                df_post_w = pd.concat([
                    df_post_w, 
                    pd.DataFrame([fila_ploss]), 
                    pd.DataFrame([fila_carga])
                ], ignore_index=True)
            
            df_post_w.to_excel(writer, sheet_name=nombre_hoja, startrow=fila_actual, startcol=0, index=False)
            fila_actual += len(df_post_w) + 2
            
            # ========================================
            # TABLA: ENS DETALLADO POR BARRA (w)
            # ========================================
            if p_ens_pre is not None and p_ens_post is not None:
                # Verificar si hay corte de carga
                total_ens_pre = sum(p_ens_pre.X[b, w] for b in range(nb))
                total_ens_post = sum(p_ens_post.X[b, c, w] for c in range(K) for b in range(nb))
                
                if total_ens_pre > 0.001 or total_ens_post > 0.001:
                    pd.DataFrame(["ENS por barra"]).to_excel(
                        writer, sheet_name=nombre_hoja, startrow=fila_actual-1, 
                        startcol=0, index=False, header=False
                    )
                    
                    ens_data = []
                    for b in range(nb):
                        if (p_ens_pre.X[b, w] > 0.001) or any(p_ens_post.X[b, c, w] > 0.001 for c in range(K)):
                            row = {
                                'Bus': b+1,
                                'ENS pre (MW)': round(float(p_ens_pre.X[b, w]), 3)
                            }
                            for c in range(K):
                                tipo, idx = contingencias[c]
                                #nombre_col = f'Cont{c+1}: {tipo}{idx}'
                                if tipo == "gen":
                                    nombre_col = f"Cont{c+1}: gen. {idx}"
                                else:
                                    if idx == 2:
                                        nombre_col = f"Cont{c+1}: alm. 2"
                                    elif idx == 4:
                                        nombre_col = f"Cont{c+1}: alm. 4"
                                    else:
                                        nombre_col = f"Cont{c+1}: alm. ENAP"
                                row[nombre_col] = round(float(p_ens_post.X[b, c, w]), 3)
                            ens_data.append(row)
                    
                    # Fila totales ENS
                    fila_total_ens = {'Bus': 'Total', 'ENS pre (MW)': round(float(total_ens_pre), 3)}
                    for c in range(K):
                        tipo, idx = contingencias[c]
                        #nombre_col = f'Cont{c+1}: {tipo}{idx}'
                        if tipo == "gen":
                            nombre_col = f"Cont{c+1}: gen. {idx}"
                        else:
                            if idx == 2:
                                nombre_col = f"Cont{c+1}: alm. 2"
                            elif idx == 4:
                                nombre_col = f"Cont{c+1}: alm. 4"
                            else:
                                nombre_col = f"Cont{c+1}: alm. ENAP"
                        total_c = sum(p_ens_post.X[b, c, w] for b in range(nb))
                        fila_total_ens[nombre_col] = round(float(total_c), 3)
                    ens_data.append(fila_total_ens)
                    
                    df_ens_w = pd.DataFrame(ens_data)
                    df_ens_w.to_excel(writer, sheet_name=nombre_hoja, startrow=fila_actual, startcol=0, index=False)
                    fila_actual += len(df_ens_w) + 2
            
            # ========================================
            # TABLA: FLUJOS PRECONTINGENCIA (w)
            # ========================================
            if f_pre is not None and branch_from is not None and branch_to is not None:
                pd.DataFrame(["Flujos precontingencia"]).to_excel(
                    writer, sheet_name=nombre_hoja, startrow=fila_actual-1, 
                    startcol=0, index=False, header=False
                )
                
                # Construir datos: primero líneas, luego trafos
                datos_flujo_pre = []
                
                # LÍNEAS
                if lineas_idx:
                    datos_flujo_pre.append({
                        'Branch': 'LÍNEAS',
                        'Flujo (MW)': '',
                        'Cargabilidad (%)': ''
                    })
                    
                    for l in lineas_idx:
                        bus_from = int(branch_from[l])
                        bus_to = int(branch_to[l])
                        flujo = round(float(f_pre.X[l, w]), 3)
                        fm = float(vars_case['FM'][l])
                        cargabilidad = round(abs(flujo) / fm * 100, 2) if fm > 0 else 0
                        
                        # Crear el nombre del branch e incluir la descripción si aplica
                        nombre_branch = f'{bus_from}-{bus_to}'
                        if (bus_from, bus_to) in descripciones_lineas:
                            nombre_branch += f" {descripciones_lineas[(bus_from, bus_to)]}"

                        datos_flujo_pre.append({
                            'Branch': nombre_branch,
                            'Flujo (MW)': flujo,
                            'Cargabilidad (%)': cargabilidad
                        })
                
                # TRANSFORMADORES
                if trafos_idx:
                    datos_flujo_pre.append({
                        'Branch': 'TRANSFORMADORES',
                        'Flujo (MW)': '',
                        'Cargabilidad (%)': ''
                    })
                    
                    for l in trafos_idx:
                        bus_from = int(branch_from[l])
                        bus_to = int(branch_to[l])
                        flujo = round(float(f_pre.X[l, w]), 3)
                        fm = float(vars_case['FM'][l])
                        cargabilidad = round(abs(flujo) / fm * 100, 2) if fm > 0 else 0
                        
                        datos_flujo_pre.append({
                            'Branch': f'{bus_from}-{bus_to}',
                            'Flujo (MW)': flujo,
                            'Cargabilidad (%)': cargabilidad
                        })
                
                df_flujo_pre_w = pd.DataFrame(datos_flujo_pre)
                df_flujo_pre_w.to_excel(writer, sheet_name=nombre_hoja, startrow=fila_actual, startcol=0, index=False)
                fila_actual += len(df_flujo_pre_w) + 2

            # ========================================
            # TABLA: FLUJOS POSTCONTINGENCIA (w)
            # ========================================
            if f_post is not None and branch_from is not None and branch_to is not None:
                pd.DataFrame(["Flujos postcontingencia"]).to_excel(
                    writer, sheet_name=nombre_hoja, startrow=fila_actual-1, 
                    startcol=0, index=False, header=False
                )
                
                # Construir TODAS las columnas primero
                columnas = ['Branch']
                for c in range(K):
                    tipo, idx = contingencias[c]
                    if tipo == "gen":
                        base = f"Cont{c+1}: gen. {idx}"
                    else:
                        if idx == 2:
                            base = f"Cont{c+1}: alm. 2"
                        elif idx == 4:
                            base = f"Cont{c+1}: alm. 4"
                        elif idx == 11:
                            base = f"Cont{c+1}: alm. 11"
                        else:
                            base = f"Cont{c+1}: alm. ENAP"
                    
                    columnas.append(f"{base} - Flujo (MW)")
                    columnas.append(f"{base} - Carg (%)")
                
                # Construir datos con todas las contingencias
                datos_flujo_post = []
                
                # LÍNEAS
                if lineas_idx:
                    # Fila separadora con TODAS las columnas vacías
                    fila_sep_lineas = {col: '' for col in columnas}
                    fila_sep_lineas['Branch'] = 'LÍNEAS'
                    datos_flujo_post.append(fila_sep_lineas)
                    
                    # Datos de líneas
                    for l in lineas_idx:
                        bus_from = int(branch_from[l])
                        bus_to = int(branch_to[l])
                        fm = float(vars_case['FM'][l])
                        
                        # Crear el nombre del branch e incluir la descripción si aplica
                        nombre_branch = f'{bus_from}-{bus_to}'
                        if (bus_from, bus_to) in descripciones_lineas:
                            nombre_branch += f" {descripciones_lineas[(bus_from, bus_to)]}"
                            
                        fila = {'Branch': nombre_branch}
                        
                        for c in range(K):
                            tipo, idx = contingencias[c]
                            flujo = round(float(f_post.X[l, c, w]), 3)
                            cargabilidad = round(abs(flujo) / fm * 100, 2) if fm > 0 else 0
                            
                            if tipo == "gen":
                                base = f"Cont{c+1}: gen. {idx}"
                            else:
                                if idx == 2:
                                    base = f"Cont{c+1}: alm. 2"
                                elif idx == 4:
                                    base = f"Cont{c+1}: alm. 4"
                                elif idx == 11:
                                    base = f"Cont{c+1}: alm. 11"
                                else:
                                    base = f"Cont{c+1}: alm. ENAP"
                            
                            fila[f"{base} - Flujo (MW)"] = flujo
                            fila[f"{base} - Carg (%)"] = cargabilidad
                        
                        datos_flujo_post.append(fila)
                
                # TRANSFORMADORES
                if trafos_idx:
                    # Fila separadora con TODAS las columnas vacías
                    fila_sep_trafos = {col: '' for col in columnas}
                    fila_sep_trafos['Branch'] = 'TRANSFORMADORES'
                    datos_flujo_post.append(fila_sep_trafos)
                    
                    # Datos de trafos
                    for l in trafos_idx:
                        bus_from = int(branch_from[l])
                        bus_to = int(branch_to[l])
                        fm = float(vars_case['FM'][l])
                        
                        fila = {'Branch': f'{bus_from}-{bus_to}'}
                        
                        for c in range(K):
                            tipo, idx = contingencias[c]
                            flujo = round(float(f_post.X[l, c, w]), 3)
                            cargabilidad = round(abs(flujo) / fm * 100, 2) if fm > 0 else 0
                            
                            if tipo == "gen":
                                base = f"Cont{c+1}: gen. {idx}"
                            else:
                                if idx == 2:
                                    base = f"Cont{c+1}: alm. 2"
                                elif idx == 4:
                                    base = f"Cont{c+1}: alm. 4"
                                elif idx == 11:
                                    base = f"Cont{c+1}: alm. 11"
                                else:
                                    base = f"Cont{c+1}: alm. ENAP"
                            
                            fila[f"{base} - Flujo (MW)"] = flujo
                            fila[f"{base} - Carg (%)"] = cargabilidad
                        
                        datos_flujo_post.append(fila)
                
                # Crear DataFrame con columnas ordenadas
                df_flujo_post_w = pd.DataFrame(datos_flujo_post, columns=columnas)
                df_flujo_post_w.to_excel(writer, sheet_name=nombre_hoja, startrow=fila_actual, startcol=0, index=False)
                
                fila_actual += len(df_flujo_post_w) + 3

# 4. Formateo Final
wb = load_workbook(nombre_archivo)

from openpyxl.styles import Alignment, Font

for hoja in wb.sheetnames:
    ws = wb[hoja]
    
    # Encontrar todas las filas a procesar
    filas_a_procesar = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=1), start=1):
        if row[0].value and "Flujos postcontingencia" in str(row[0].value):
            filas_a_procesar.append(row_idx + 1)  # header está 1 fila después
    
    # Procesar de ARRIBA hacia ABAJO (Top-Down) usando un offset
    offset = 0
    for orig_header_row_idx in filas_a_procesar:
        # El índice real se desplaza hacia abajo por cada fila que insertamos antes
        header_row_idx = orig_header_row_idx + offset
        
        # Insertar fila para encabezados principales
        ws.insert_rows(header_row_idx)
        
        col_idx = 2  # Columna B (después de "Branch")
        while col_idx <= ws.max_column:
            cell = ws.cell(row=header_row_idx + 1, column=col_idx)
            if cell.value and "- Flujo (MW)" in str(cell.value):
                # Extraer nombre base
                base_name = str(cell.value).replace(" - Flujo (MW)", "")
                
                # Combinar celdas para el encabezado principal
                ws.merge_cells(
                    start_row=header_row_idx, start_column=col_idx,
                    end_row=header_row_idx, end_column=col_idx + 1
                )
                
                # Escribir nombre de contingencia en la celda combinada
                merged_cell = ws.cell(row=header_row_idx, column=col_idx)
                merged_cell.value = base_name
                merged_cell.alignment = Alignment(horizontal='center', vertical='center')
                merged_cell.font = Font(bold=True)
                
                # Actualizar subencabezados
                ws.cell(row=header_row_idx + 1, column=col_idx).value = "Flujo (MW)"
                ws.cell(row=header_row_idx + 1, column=col_idx + 1).value = "Carg (%)"
                
                col_idx += 2
            else:
                col_idx += 1
        
        # Incrementar el offset porque acabamos de insertar 1 fila
        offset += 1
    
    # Ajustar ancho de columnas AL FINAL
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 30)

wb.save(nombre_archivo)
print(f"\nArchivo '{nombre_archivo}' generado exitosamente con ENS, flujos separados y cargabilidad.")

#os.startfile(nombre_archivo)

# ============================================================================
# Gráficos combinados (pre+post) por contingencia objetivo, y gráfico de
# reserva/vertimiento del Parque Eólico (PE) — exportación a PDF
# ----------------------------------------------------------------------------
# Pegar después de que `vars_list` esté construido y el modelo resuelto
# (después de m.optimize()). No depende de la exportación a Excel.
# Reemplaza a graficos_barras_main.py de la versión anterior.
# ============================================================================

import os
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.patches import Patch

# ============================================================================
# 1. CONFIGURACIÓN DEL USUARIO
# ============================================================================

FIGSIZE = (9, 5.5)   # ancho, alto [pulgadas] -> tamaño físico real del PDF (AJUSTAR PARA EL INFORME)
DPI = 300            # poco relevante en PDF (vectorial); útil si exportas también a PNG

CARPETA_SALIDA = "graficos_resultados"

CASOS_A_GRAFICAR = 'all'          # 'all' o lista, ej: [1, 2]
TIPOS_A_GRAFICAR = 'all'          # 'all' o subconjunto de {'despacho', 'flujos', 'pe'}

ORDENAR_POR_ETA = True            # eje x ordenado por valor de eta (ascendente)

# --- Contingencias post-contingencia de interés (búsqueda dinámica por caso) ---
CONTINGENCIAS_OBJETIVO = [
    {'tipo': 'gen',  'busqueda': 'Solar Titan', 'slug': 'solar_titan'},
    {'tipo': 'load', 'numero': 2, 'etiqueta': 'Salida Alimentador 2', 'slug': 'alimentador_2'},
]

# --- Ramas a incluir en los gráficos de flujos (todo lo demás se filtra) ---
# Nota: se corrigió una coma faltante y una entrada (5,6)->(5,6) duplicada,
# que ahora es (5,6)->(6,5) para mantener ambos sentidos como el resto.
descripciones_flujos = {
    # Líneas
    (4, 5): "(66 kV)", (5, 4): "(66 kV)",
    (7, 8): "(PE)", (8, 7): "(PE)",
    (8, 9): "(PE)", (9, 8): "(PE)",
    (9, 10): "(PE)", (10, 9): "(PE)",
    # Trafos
    (1, 2): "(T6)", (2, 1): "(T6)",
    (1, 3): "(T2)", (3, 1): "(T2)",
    (1, 7): "(T1)", (7, 1): "(T1)",
    (1, 4): "(T5 y T7)", (4, 1): "(T5 y T7)",
    (5, 6): "(T7)", (6, 5): "(T7)",
}

# ============================================================================
# 2. PALETA DE COLORES
# ============================================================================

# Térmicas / otras unidades no-ERV (sin cambios). Se cicla con % si faltan colores.
colores_termicos = ['#7C4B3A', '#7D6642', '#B8643B', '#B78D43', '#dd8452', '#C44E52']

# WTG eólicos: tono OSCURO base por unidad (la parte clara se deriva con lighten_color)
colores_wtg = ['#2E7D32', '#3C8F5C', '#6FAE8B']

# Líneas/transformadores: paleta ampliada y diversa (antes solo azul/morado dificultaba
# la lectura con muchas ramas). Se cicla con % si faltan colores.
colores_flujos_pool = [
    '#4C72B0', '#8172B3', '#36A2A6', '#C2548F', '#5B5EA6',
    '#5C7A89', '#8E4585', '#4F6D7A', '#A0527C', '#3E5C76',
]

# Paleta específica del gráfico PE (reserva / vertimiento)
VERDE_PE = '#2E7D32'
NARANJA_PE = '#E08214'
AZUL_PE = '#1F77B4'
AMARILLO_PE = '#FDD835'

def lighten_color(hex_color, factor=0.55):
    """Mezcla un color hex con blanco; factor=0 -> sin cambio, 1 -> blanco."""
    hex_color = hex_color.lstrip('#')
    r, g, b = (int(hex_color[i:i+2], 16) for i in (0, 2, 4))
    r = int(r + (255 - r) * factor)
    g = int(g + (255 - g) * factor)
    b = int(b + (255 - b) * factor)
    return f'#{r:02x}{g:02x}{b:02x}'

def color_generador(idx, idx_erv):
    if idx in idx_erv:
        pos = list(idx_erv).index(idx)
        return colores_wtg[pos % len(colores_wtg)]
    return colores_termicos[idx % len(colores_termicos)]

def color_linea(idx):
    return colores_flujos_pool[idx % len(colores_flujos_pool)]

# ============================================================================
# 3. UTILIDADES
# ============================================================================

os.makedirs(CARPETA_SALIDA, exist_ok=True)
contador_figuras = 0

def nombre_limpio(g_name):
    return g_name.replace('.ElmSym', '').replace('.ElmGenstat', '')

def tipo_incluido(tipo):
    return TIPOS_A_GRAFICAR == 'all' or tipo in TIPOS_A_GRAFICAR

def guardar_figura(fig, nombre_archivo):
    global contador_figuras
    ruta = os.path.join(CARPETA_SALIDA, nombre_archivo)
    fig.savefig(ruta, format='pdf', dpi=DPI, bbox_inches='tight')
    plt.close(fig)
    contador_figuras += 1
    print(f'  [{contador_figuras}] {nombre_archivo}')

def buscar_contingencia(vars_case, objetivo):
    """
    Devuelve (k_idx, etiqueta) de la contingencia objetivo en este caso,
    o (None, None) si no existe (ej. generador no activo en este caso).
    """
    contingencias = vars_case['contingencias']
    gen_names = vars_case['gen_names']

    if objetivo['tipo'] == 'gen':
        idx_gen = next(
            (i for i, n in enumerate(gen_names) if objetivo['busqueda'] in n), None
        )
        if idx_gen is None:
            return None, None
        try:
            k_idx = contingencias.index(('gen', idx_gen + 1))
        except ValueError:
            return None, None
        etiqueta = objetivo.get('etiqueta') or f"Salida {nombre_limpio(gen_names[idx_gen])}"
        return k_idx, etiqueta

    elif objetivo['tipo'] == 'load':
        try:
            k_idx = contingencias.index(('load', objetivo['numero']))
        except ValueError:
            return None, None
        return k_idx, objetivo.get('etiqueta', f"Salida carga {objetivo['numero']}")

    return None, None

def construir_eta_matrix(ng, idx_erv, eta_ordenado, gen_fuera_servicio=None):
    """Matriz (ng, n_w) con eta solo en las filas WTG; 0 en la fila que esté fuera de servicio."""
    eta_m = np.zeros((ng, len(eta_ordenado)))
    for i in idx_erv:
        if i == gen_fuera_servicio:
            continue
        eta_m[i, :] = eta_ordenado
    return eta_m

def dibujar_stack_despacho(ax, x_array, ancho, p_matrix, eta_matrix, ng, idx_erv,
                            colores_gen, hatch=None):
    """
    Dibuja un stacked bar por elemento de x_array. Para generadores WTG, separa
    el bloque en despachable (oscuro) + estocástico (claro) con línea punteada
    en el límite. Devuelve el array `bottom` final (tope de cada barra).
    """
    bottom = np.zeros(len(x_array))
    for i in range(ng):
        if i in idx_erv:
            color_dark = colores_gen[i]
            color_light = lighten_color(color_dark, 0.55)

            y_desp = p_matrix[i, :]
            ax.bar(x_array, y_desp, ancho, bottom=bottom, color=color_dark,
                   edgecolor='black', linewidth=0.5, hatch=hatch)
            bottom = bottom + y_desp

            y_esto = eta_matrix[i, :]
            ax.bar(x_array, y_esto, ancho, bottom=bottom, color=color_light,
                   edgecolor='black', linewidth=0.5, hatch=hatch)
            for xc, b in zip(x_array, bottom):
                ax.plot([xc - ancho/2, xc + ancho/2], [b, b], linestyle='--',
                        color='black', linewidth=0.6, zorder=5)
            bottom = bottom + y_esto
        else:
            y = p_matrix[i, :]
            ax.bar(x_array, y, ancho, bottom=bottom, color=colores_gen[i],
                   edgecolor='black', linewidth=0.5, hatch=hatch)
            bottom = bottom + y
    return bottom

# ============================================================================
# 4. GENERACIÓN DE GRÁFICOS POR CASO
# ============================================================================

if status == GRB.Status.OPTIMAL:

    sort_idx = np.argsort(eta_list) if ORDENAR_POR_ETA else np.arange(n_w)
    eta_ordenado_global = eta_list[sort_idx]
    etiquetas_x = [f'{v:.2f}' for v in eta_ordenado_global]
    x_pos = np.arange(n_w)

    casos_a_iterar = casos_ejecutar if CASOS_A_GRAFICAR == 'all' else CASOS_A_GRAFICAR

    for u_idx, u in enumerate(casos_ejecutar):
        if u not in casos_a_iterar:
            continue

        vars_case = vars_list[u_idx]

        p_pre_val  = vars_case['p_pre'].X
        f_pre_val  = vars_case['f_pre'].X
        p_post_val = vars_case['p_post'].X
        f_post_val = vars_case['f_post'].X

        gen_names     = vars_case['gen_names']
        branch_from   = vars_case['branch_from']
        branch_to     = vars_case['branch_to']
        contingencias = vars_case['contingencias']
        FM_caso       = vars_case['FM']

        ng = p_pre_val.shape[0]
        nl = f_pre_val.shape[0]
        idx_erv = list(range(ng - 3, ng))  # últimos 3 = WTG (convención del modelo)

        gen_labels = [nombre_limpio(g) for g in gen_names]
        colores_gen_caso = [color_generador(i, idx_erv) for i in range(ng)]

        p_pre_ord = p_pre_val[:, sort_idx]
        f_pre_ord = f_pre_val[:, sort_idx]

        print(f'\nCaso {u}:')

        # ------------------------------------------------------------------
        # 4.1 DESPACHO COMBINADO (pre + post) por contingencia objetivo
        # ------------------------------------------------------------------
        if tipo_incluido('despacho'):
            for objetivo in CONTINGENCIAS_OBJETIVO:
                k_idx, etiqueta = buscar_contingencia(vars_case, objetivo)
                if k_idx is None:
                    ref = objetivo.get('busqueda', objetivo.get('numero'))
                    print(f'  [Aviso] Contingencia "{ref}" no existe en el Caso {u}; se omite (despacho).')
                    continue

                tipo_k, idx_k = contingencias[k_idx]
                gen_fs = (idx_k - 1) if (tipo_k == 'gen' and (idx_k - 1) in idx_erv) else None

                eta_matrix_pre  = construir_eta_matrix(ng, idx_erv, eta_ordenado_global, gen_fuera_servicio=None)
                eta_matrix_post = construir_eta_matrix(ng, idx_erv, eta_ordenado_global, gen_fuera_servicio=gen_fs)
                p_post_ord = p_post_val[:, k_idx, sort_idx]

                fig, ax = plt.subplots(figsize=FIGSIZE)
                ancho = 0.32
                x_pre = x_pos - ancho/2 - 0.02
                x_post = x_pos + ancho/2 + 0.02

                dibujar_stack_despacho(ax, x_pre, ancho, p_pre_ord, eta_matrix_pre,
                                        ng, idx_erv, colores_gen_caso, hatch=None)
                dibujar_stack_despacho(ax, x_post, ancho, p_post_ord, eta_matrix_post,
                                        ng, idx_erv, colores_gen_caso, hatch='///')

                ax.set_xticks(x_pos); ax.set_xticklabels(etiquetas_x, rotation=45, ha='right')
                ax.set_xlabel('Error estocástico η [MW]')
                ax.set_ylabel('Potencia [MW]')
                ax.set_title(f'Caso {u} — Despachos pre/post ({etiqueta})')

                handles = [Patch(facecolor=colores_gen_caso[i], edgecolor='black', label=gen_labels[i])
                           for i in range(ng) if i not in idx_erv]
                handles += [Patch(facecolor=colores_wtg[0], edgecolor='black', label='WTG (desp.)'),
                            Patch(facecolor=lighten_color(colores_wtg[0], 0.55), edgecolor='black', label='WTG (estoc.)')]
                handles += [Patch(facecolor='white', edgecolor='black', label='Pre'),
                            Patch(facecolor='white', edgecolor='black', hatch='///', label='Post')]
                ax.legend(handles=handles, loc='upper center', bbox_to_anchor=(0.5, -0.30), ncol=4, fontsize=7)
                ax.grid(True, axis='y', alpha=0.3)
                plt.tight_layout()
                guardar_figura(fig, f'caso{u}_despacho_{objetivo["slug"]}.pdf')

        # ------------------------------------------------------------------
        # 4.2 FLUJOS COMBINADOS (cargabilidad %, pre + post) por contingencia
        # ------------------------------------------------------------------
        if tipo_incluido('flujos'):
            idx_mostrar = [
                l for l in range(nl)
                if (int(branch_from[l]), int(branch_to[l])) in descripciones_flujos
                or (int(branch_to[l]), int(branch_from[l])) in descripciones_flujos
            ]
            if not idx_mostrar:
                print(f'  [Aviso] Ninguna rama del Caso {u} coincide con descripciones_flujos; se omiten flujos.')
            else:
                line_labels = []
                for l in idx_mostrar:
                    bf, bt = int(branch_from[l]), int(branch_to[l])
                    desc = descripciones_flujos.get((bf, bt), descripciones_flujos.get((bt, bf), ''))
                    line_labels.append(f'{bf}-{bt} {desc}'.strip())

                for objetivo in CONTINGENCIAS_OBJETIVO:
                    k_idx, etiqueta = buscar_contingencia(vars_case, objetivo)
                    if k_idx is None:
                        ref = objetivo.get('busqueda', objetivo.get('numero'))
                        print(f'  [Aviso] Contingencia "{ref}" no existe en el Caso {u}; se omite (flujos).')
                        continue

                    fig, ax = plt.subplots(figsize=FIGSIZE)
                    n_l = len(idx_mostrar)
                    ancho_grupo = 0.85
                    ancho_barra = ancho_grupo / (n_l * 2)

                    for j, l in enumerate(idx_mostrar):
                        color_pre = color_linea(j)
                        color_post = lighten_color(color_pre, 0.5)
                        off_pre  = -ancho_grupo/2 + (2*j + 0.5) * ancho_barra
                        off_post = -ancho_grupo/2 + (2*j + 1.5) * ancho_barra
                        fm = FM_caso[l]
                        carg_pre  = np.abs(f_pre_ord[l, :]) / fm * 100
                        carg_post = np.abs(f_post_val[l, k_idx, sort_idx]) / fm * 100
                        ax.bar(x_pos + off_pre, carg_pre, ancho_barra, color=color_pre,
                               edgecolor='black', linewidth=0.4)
                        ax.bar(x_pos + off_post, carg_post, ancho_barra, color=color_post,
                               edgecolor='black', linewidth=0.4)

                    ax.set_xticks(x_pos); ax.set_xticklabels(etiquetas_x, rotation=45, ha='right')
                    ax.set_xlabel('Error estocástico η [MW]')
                    ax.set_ylabel('Cargabilidad [%]')
                    ax.set_title(f'Caso {u} — Cargabilidad de ramas, pre/post ({etiqueta})')

                    handles = [Patch(facecolor=color_linea(j), edgecolor='black', label=line_labels[j])
                               for j in range(n_l)]
                    handles += [Patch(facecolor='dimgray', edgecolor='black', label='Pre (oscuro)'),
                                Patch(facecolor=lighten_color('#444444', 0.55), edgecolor='black', label='Post (claro)')]
                    ax.legend(handles=handles, loc='upper center', bbox_to_anchor=(0.5, -0.34), ncol=4, fontsize=7)
                    ax.grid(True, axis='y', alpha=0.3)
                    plt.tight_layout()
                    guardar_figura(fig, f'caso{u}_flujos_{objetivo["slug"]}.pdf')

        # ------------------------------------------------------------------
        # 4.3 PARQUE EÓLICO (PE): reserva y vertimiento — solo precontingencia
        # ------------------------------------------------------------------
        if tipo_incluido('pe'):
            # r_up es por caso pero no se guarda en vars_case; se reconstruye desde
            # r_up_g (global) con el mismo mapeo que ya usa el modelo (dicc_gen).
            idx_activos_caso = [dicc_gen[name] for name in gen_names]
            r_up_caso = r_up_g.X[idx_activos_caso]

            n_erv = len(idx_erv)
            p_pre_total  = p_pre_ord[idx_erv, :].sum(axis=0)
            eta_total    = n_erv * eta_ordenado_global
            r_up_total   = r_up_caso[idx_erv].sum()
            p_fore_total = n_erv * p_fore
            p_VUL_total  = n_erv * p_VUL
            vertimiento  = np.maximum(0.0, p_VUL_total - p_pre_total - r_up_total)

            fig, ax = plt.subplots(figsize=FIGSIZE)
            ancho = 0.5
            bottom = np.zeros(n_w)
            ax.bar(x_pos, p_pre_total, ancho, bottom=bottom, color=VERDE_PE,
                   edgecolor='black', linewidth=0.5, label='Despachable (Σ p_pre)')
            bottom = bottom + p_pre_total
            ax.bar(x_pos, eta_total, ancho, bottom=bottom, color=NARANJA_PE,
                   edgecolor='black', linewidth=0.5, label='Estocástico (Σ η)')
            bottom = bottom + eta_total
            r_up_arr = np.full(n_w, r_up_total)
            ax.bar(x_pos, r_up_arr, ancho, bottom=bottom, color=AZUL_PE,
                   edgecolor='black', linewidth=0.5, label='Reserva de subida (Σ r_up)')
            bottom = bottom + r_up_arr
            ax.bar(x_pos, vertimiento, ancho, bottom=bottom, color=AMARILLO_PE,
                   edgecolor='black', linewidth=0.5, label='Vertimiento no utilizable')

            ax.axhline(p_fore_total, color='black', linestyle='--', linewidth=1.2,
                       label=f'P_fore total = {p_fore_total:.2f} MW')
            ax.axhline(p_VUL_total, color='dimgray', linestyle=':', linewidth=1.2,
                       label=f'P_VUL total = {p_VUL_total:.2f} MW')

            ax.set_xticks(x_pos); ax.set_xticklabels(etiquetas_x, rotation=45, ha='right')
            ax.set_xlabel('Error estocástico η [MW]')
            ax.set_ylabel('Potencia [MW]')
            ax.set_title(f'Caso {u} — Parque Eólico (PE): reserva y vertimiento (pre-contingencia)')
            ax.legend(loc='upper center', bbox_to_anchor=(0.5, -0.28), ncol=2, fontsize=8)
            ax.grid(True, axis='y', alpha=0.3)
            plt.tight_layout()
            guardar_figura(fig, f'caso{u}_pe_reserva.pdf')

    print(f'\nTotal: {contador_figuras} gráficos exportados en {os.path.abspath(CARPETA_SALIDA)}')

else:
    print('No se generan gráficos: el modelo no alcanzó óptimo (status %d).' % status)
