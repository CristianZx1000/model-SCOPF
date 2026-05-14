if not status == GRB.Status.OPTIMAL:
    ejecutar = False
    assert ejecutar, "Ejecución detenida por condición"

import pandas as pd
from openpyxl import load_workbook
import numpy as np
import os

nombre_archivo = "resultados_SM_PA_v7_estocastico.xlsx"

# 1. Preparar DataFrame de reservas globales
datos_reservas = []
for name in gen_agc:
    matches = [key for key in dicc_gen.keys() if name in key]
    i = dicc_gen[matches[0]]
    datos_reservas.append({
        "Generador": name,
        "Reserva Up (MW)": round(float(r_up_g.X[i]), 3),
        "Reserva Dn (MW)": round(float(r_dn_g.X[i]), 3)
    })
df_reservas = pd.DataFrame(datos_reservas)

# ExcelWriter
with pd.ExcelWriter(nombre_archivo, engine="openpyxl") as writer:
    
    # 2. Iterar sobre cada CASO operativo (u)
    for u_idx, vars_case in enumerate(vars_list):
        nombre_hoja = f"Caso {u_idx+1}"
        
        # Recuperar variables del caso
        p_pre = vars_case['p_pre']
        ploss_pre = vars_case['ploss_pre']
        p_post = vars_case['p_post']
        ploss_post = vars_case['ploss_post']
        
        # Variables adicionales
        p_ens_pre = vars_case.get('p_ens_pre')
        p_ens_post = vars_case.get('p_ens_post')
        f_pre = vars_case.get('f_pre')
        f_post = vars_case.get('f_post')
        branch_from = vars_case.get('branch_from')
        branch_to = vars_case.get('branch_to')
        
        # Convertir costos a float
        try:
            val_pre = vars_case['Cop_pre'].getValue()
            val_post = vars_case['Cop_post'].getValue()
            Cop_pre_val = float(val_pre) if np.ndim(val_pre) == 0 else float(val_pre.item())
            Cop_post_val = float(val_post) if np.ndim(val_post) == 0 else float(val_post.item())
        except:
            Cop_pre_val = float(vars_case['Cop_pre'].getValue())
            Cop_post_val = float(vars_case['Cop_post'].getValue())
        
        Load_bus_pre = vars_case['Load_bus_pre']
        Load_bus_post = vars_case['Load_bus_post']
        
        gen_names = vars_case['gen_names']
        contingencias = vars_case['contingencias']
        
        ng = p_pre.shape[0]
        n_w_caso = p_pre.shape[1]
        K = p_post.shape[1]
        
        # Obtener dimensiones si existen las variables
        nb = p_ens_pre.shape[0] if p_ens_pre is not None else 0
        nl = f_pre.shape[0] if f_pre is not None else 0

        # --- ESCRIBIR ENCABEZADOS Y RESERVAS ---
        fila_actual = 0
        
        # Título del caso
        pd.DataFrame([f"Resultados {nombre_hoja}"]).to_excel(
            writer, sheet_name=nombre_hoja, startrow=fila_actual, 
            startcol=0, index=False, header=False
        )
        fila_actual += 1
        
        # Tabla de costos esperados
        df_costos = pd.DataFrame([
            #{"Parámetro": "Costo total (valor esperado) ($/h)", "Valor": round(Cop_pre_val + Cop_post_val, 2)},
            {"Parámetro": "Costo total (valor esperado) ($/h)", "Valor": round(m.objVal, 2)},
            {"Parámetro": "Costo precontingencia (valor esperado)", "Valor": round(Cop_pre_val, 2)},
            {"Parámetro": "Costo reserva", "Valor": round(float(C_res.getValue()), 2)},
            {"Parámetro": "Costo postcontingencia (valor esperado)", "Valor": round(Cop_post_val, 2)},
            {"Parámetro": "Pronóstico (p_fore) [MW]", "Valor": round(p_fore, 2)},
            {"Parámetro": "Límite VUL (p_VUL) [MW]", "Valor": round(p_VUL, 2)}
        ])
        df_costos.to_excel(writer, sheet_name=nombre_hoja, startrow=fila_actual, startcol=0, index=False)
        
        # Tabla de reservas
        df_reservas.to_excel(writer, sheet_name=nombre_hoja, startrow=fila_actual, startcol=4, index=False)
        
        fila_actual += max(len(df_costos), len(df_reservas)) + 2

        # 3. Iterar sobre cada escenario de incertidumbre (w)
        for w in range(n_w_caso):
            epsilon_val = epsilon_list[w]
            eta_val = eta_list[w]
            
            # Encabezado escenario incertidumbre
            header_w = f"Escenario {w+1} (incertidumbre: {epsilon_val:.3f} MW | eta: {eta_val:.3f} MW)"
            # header_w = f"Escenario de incertidumbre {w+1} (Pronóstico: {p_fore:.2f} MW | p_VUL: {p_VUL:.2f} MW)"
            pd.DataFrame([header_w]).to_excel(
                writer, sheet_name=nombre_hoja, startrow=fila_actual, 
                startcol=0, index=False, header=False
            )
            fila_actual += 1
            
            # ========================================
            # TABLA: PRECONTINGENCIA (w)
            # ========================================
            datos_pre = []
            for h in range(ng):
                clean_name = gen_names[h].replace('.ElmSym', '').replace('.ElmGenstat', '')
                datos_pre.append({
                    "Generador": clean_name,
                    "Potencia pre (MW)": round(float(p_pre.X[h, w]), 3)
                })
            
            datos_pre.append({
                "Generador": "Pérdidas totales",
                "Potencia pre (MW)": round(float(ploss_pre.X[:, w].sum()), 3)
            })
            
            # ENS precontingencia
            if p_ens_pre is not None:
                ens_pre_total = sum(p_ens_pre.X[b, w] for b in range(nb))
                datos_pre.append({
                    "Generador": "ENS total",
                    "Potencia pre (MW)": round(float(ens_pre_total), 3)
                })
            
            datos_pre.append({
                "Generador": "Carga total",
                "Potencia pre (MW)": round(float(sum(Load_bus_pre)), 3)
            })
            
            df_pre_w = pd.DataFrame(datos_pre)
            df_pre_w.to_excel(writer, sheet_name=nombre_hoja, startrow=fila_actual, startcol=0, index=False)
            fila_actual += len(df_pre_w) + 2
            
            # ========================================
            # TABLA: POSTCONTINGENCIA (w)
            # ========================================
            pd.DataFrame(["Postcontingencia"]).to_excel(
                writer, sheet_name=nombre_hoja, startrow=fila_actual-1, 
                startcol=0, index=False, header=False
            )
            
            datos_post_dict = {}
            lista_generadores = [gen_names[h].replace('.ElmSym', '').replace('.ElmGenstat', '') for h in range(ng)]
            datos_post_dict["Generador\ contingencia"] = lista_generadores
            
            for c in range(K):
                tipo, idx = contingencias[c]
                #nombre_col = f"Cont{c+1}: {tipo}{idx}"
                if tipo == "gen":
                    nombre_col = f"Cont{c+1}: gen. {idx}"
                else:
                    if idx == 2:
                        nombre_col = f"Cont{c+1}: alm. 2"
                    elif idx == 4:
                        nombre_col = f"Cont{c+1}: alm. 4"
                    else:
                        nombre_col = f"Cont{c+1}: alm. ENAP"
                datos_post_dict[nombre_col] = [round(float(p_post.X[h, c, w]), 3) for h in range(ng)]
            
            df_post_w = pd.DataFrame(datos_post_dict)
            
            # Fila de pérdidas
            fila_ploss = {"Generador\ contingencia": "Pérdidas (MW)"}
            for c in range(K):
                tipo, idx = contingencias[c]
                #nombre_col = f"Cont{c+1}: {tipo}{idx}"
                if tipo == "gen":
                    nombre_col = f"Cont{c+1}: gen. {idx}"
                else:
                    if idx == 2:
                        nombre_col = f"Cont{c+1}: alm. 2"
                    elif idx == 4:
                        nombre_col = f"Cont{c+1}: alm. 4"
                    else:
                        nombre_col = f"Cont{c+1}: alm. ENAP"
                fila_ploss[nombre_col] = round(float(ploss_post.X[:, c, w].sum()), 3)
            
            # Fila ENS total
            if p_ens_post is not None:
                fila_ens = {"Generador\ contingencia": "ENS total (MW)"}
                for c in range(K):
                    tipo, idx = contingencias[c]
                    #nombre_col = f"Cont{c+1}: {tipo}{idx}"
                    if tipo == "gen":
                        nombre_col = f"Cont{c+1}: gen. {idx}"
                    else:
                        if idx == 2:
                            nombre_col = f"Cont{c+1}: alm. 2"
                        elif idx == 4:
                            nombre_col = f"Cont{c+1}: alm. 4"
                        else:
                            nombre_col = f"Cont{c+1}: alm. ENAP"
                    ens_total = p_ens_post.X[:, c, w].sum()
                    fila_ens[nombre_col] = round(float(ens_total), 3)
            
            # Fila de carga
            fila_carga = {"Generador\ contingencia": "Carga total (MW)"}
            for c in range(K):
                tipo, idx = contingencias[c]
                nombre_k = f"Cont{c+1}_{tipo}{idx}"
                Load_bus_k = Load_bus_post[nombre_k]
                #nombre_col = f"Cont{c+1}: {tipo}{idx}"
                if tipo == "gen":
                    nombre_col = f"Cont{c+1}: gen. {idx}"
                else:
                    if idx == 2:
                        nombre_col = f"Cont{c+1}: alm. 2"
                    elif idx == 4:
                        nombre_col = f"Cont{c+1}: alm. 4"
                    else:
                        nombre_col = f"Cont{c+1}: alm. ENAP"
                fila_carga[nombre_col] = round(float(sum(Load_bus_k)), 3)
            
            # Concatenar todas las filas
            if p_ens_post is not None:
                df_post_w = pd.concat([
                    df_post_w, 
                    pd.DataFrame([fila_ploss]), 
                    pd.DataFrame([fila_ens]),
                    pd.DataFrame([fila_carga])
                ], ignore_index=True)
            else:
                df_post_w = pd.concat([
                    df_post_w, 
                    pd.DataFrame([fila_ploss]), 
                    pd.DataFrame([fila_carga])
                ], ignore_index=True)
            
            df_post_w.to_excel(writer, sheet_name=nombre_hoja, startrow=fila_actual, startcol=0, index=False)
            fila_actual += len(df_post_w) + 2
            
            # ========================================
            # TABLA: ENS DETALLADO POR BARRA (w)
            # ========================================
            if p_ens_pre is not None and p_ens_post is not None:
                # Verificar si hay corte de carga
                total_ens_pre = sum(p_ens_pre.X[b, w] for b in range(nb))
                total_ens_post = sum(p_ens_post.X[b, c, w] for c in range(K) for b in range(nb))
                
                if total_ens_pre > 0.001 or total_ens_post > 0.001:
                    pd.DataFrame(["ENS por barra"]).to_excel(
                        writer, sheet_name=nombre_hoja, startrow=fila_actual-1, 
                        startcol=0, index=False, header=False
                    )
                    
                    ens_data = []
                    for b in range(nb):
                        if (p_ens_pre.X[b, w] > 0.001) or any(p_ens_post.X[b, c, w] > 0.001 for c in range(K)):
                            row = {
                                'Bus': b+1,
                                'ENS pre (MW)': round(float(p_ens_pre.X[b, w]), 3)
                            }
                            for c in range(K):
                                tipo, idx = contingencias[c]
                                #nombre_col = f'Cont{c+1}: {tipo}{idx}'
                                if tipo == "gen":
                                    nombre_col = f"Cont{c+1}: gen. {idx}"
                                else:
                                    if idx == 2:
                                        nombre_col = f"Cont{c+1}: alm. 2"
                                    elif idx == 4:
                                        nombre_col = f"Cont{c+1}: alm. 4"
                                    else:
                                        nombre_col = f"Cont{c+1}: alm. ENAP"
                                row[nombre_col] = round(float(p_ens_post.X[b, c, w]), 3)
                            ens_data.append(row)
                    
                    # Fila totales ENS
                    fila_total_ens = {'Bus': 'Total', 'ENS pre (MW)': round(float(total_ens_pre), 3)}
                    for c in range(K):
                        tipo, idx = contingencias[c]
                        #nombre_col = f'Cont{c+1}: {tipo}{idx}'
                        if tipo == "gen":
                            nombre_col = f"Cont{c+1}: gen. {idx}"
                        else:
                            if idx == 2:
                                nombre_col = f"Cont{c+1}: alm. 2"
                            elif idx == 4:
                                nombre_col = f"Cont{c+1}: alm. 4"
                            else:
                                nombre_col = f"Cont{c+1}: alm. ENAP"
                        total_c = sum(p_ens_post.X[b, c, w] for b in range(nb))
                        fila_total_ens[nombre_col] = round(float(total_c), 3)
                    ens_data.append(fila_total_ens)
                    
                    df_ens_w = pd.DataFrame(ens_data)
                    df_ens_w.to_excel(writer, sheet_name=nombre_hoja, startrow=fila_actual, startcol=0, index=False)
                    fila_actual += len(df_ens_w) + 2
            
            # ========================================
            # TABLA: FLUJOS PRECONTINGENCIA (w)
            # ========================================
            if f_pre is not None and branch_from is not None and branch_to is not None:
                pd.DataFrame(["Flujos precontingencia"]).to_excel(
                    writer, sheet_name=nombre_hoja, startrow=fila_actual-1, 
                    startcol=0, index=False, header=False
                )
                
                # Obtener índices de líneas y trafos desde network
                line_indices = network["line_indices"]
                
                # Identificar qué branches son líneas y cuáles trafos
                lineas_idx = []
                trafos_idx = []
                
                for l in range(nl):
                    bus_from = int(branch_from[l])
                    bus_to = int(branch_to[l])
                    
                    # Buscar si es trafo o línea
                    es_trafo = False
                    for key, info in line_indices.items():
                        if info["fbus"] == bus_from and info["tbus"] == bus_to:
                            if "trafo" in key.lower():
                                es_trafo = True
                            break
                    
                    if es_trafo:
                        trafos_idx.append(l)
                    else:
                        lineas_idx.append(l)
                
                # Construir datos: primero líneas, luego trafos
                datos_flujo_pre = []
                
                # LÍNEAS
                if lineas_idx:
                    datos_flujo_pre.append({
                        'Branch': 'LÍNEAS',
                        'Flujo (MW)': '',
                        'Cargabilidad (%)': ''
                    })
                    
                    for l in lineas_idx:
                        bus_from = int(branch_from[l])
                        bus_to = int(branch_to[l])
                        flujo = round(float(f_pre.X[l, w]), 3)
                        fm = float(vars_case['FM'][l])
                        cargabilidad = round(abs(flujo) / fm * 100, 2) if fm > 0 else 0
                        
                        datos_flujo_pre.append({
                            'Branch': f'{bus_from}-{bus_to}',
                            'Flujo (MW)': flujo,
                            'Cargabilidad (%)': cargabilidad
                        })
                
                # TRANSFORMADORES
                if trafos_idx:
                    datos_flujo_pre.append({
                        'Branch': 'TRANSFORMADORES',
                        'Flujo (MW)': '',
                        'Cargabilidad (%)': ''
                    })
                    
                    for l in trafos_idx:
                        bus_from = int(branch_from[l])
                        bus_to = int(branch_to[l])
                        flujo = round(float(f_pre.X[l, w]), 3)
                        fm = float(vars_case['FM'][l])
                        cargabilidad = round(abs(flujo) / fm * 100, 2) if fm > 0 else 0
                        
                        datos_flujo_pre.append({
                            'Branch': f'{bus_from}-{bus_to}',
                            'Flujo (MW)': flujo,
                            'Cargabilidad (%)': cargabilidad
                        })
                
                df_flujo_pre_w = pd.DataFrame(datos_flujo_pre)
                df_flujo_pre_w.to_excel(writer, sheet_name=nombre_hoja, startrow=fila_actual, startcol=0, index=False)
                fila_actual += len(df_flujo_pre_w) + 2

            # ========================================
            # TABLA: FLUJOS POSTCONTINGENCIA (w)
            # ========================================
            if f_post is not None and branch_from is not None and branch_to is not None:
                pd.DataFrame(["Flujos postcontingencia"]).to_excel(
                    writer, sheet_name=nombre_hoja, startrow=fila_actual-1, 
                    startcol=0, index=False, header=False
                )
                
                # Identificar líneas y trafos
                lineas_idx = []
                trafos_idx = []
                
                for l in range(nl):
                    bus_from = int(branch_from[l])
                    bus_to = int(branch_to[l])
                    
                    es_trafo = False
                    for key, info in line_indices.items():
                        if info["fbus"] == bus_from and info["tbus"] == bus_to:
                            if "trafo" in key.lower():
                                es_trafo = True
                            break
                    
                    if es_trafo:
                        trafos_idx.append(l)
                    else:
                        lineas_idx.append(l)
                
                # Construir TODAS las columnas primero
                columnas = ['Branch']
                for c in range(K):
                    tipo, idx = contingencias[c]
                    if tipo == "gen":
                        base = f"Cont{c+1}: gen. {idx}"
                    else:
                        if idx == 2:
                            base = f"Cont{c+1}: alm. 2"
                        elif idx == 4:
                            base = f"Cont{c+1}: alm. 4"
                        elif idx == 11:
                            base = f"Cont{c+1}: alm. 11"
                        else:
                            base = f"Cont{c+1}: alm. ENAP"
                    
                    columnas.append(f"{base} - Flujo (MW)")
                    columnas.append(f"{base} - Carg (%)")
                
                # Construir datos con todas las contingencias
                datos_flujo_post = []
                
                # LÍNEAS
                if lineas_idx:
                    # Fila separadora con TODAS las columnas vacías
                    fila_sep_lineas = {col: '' for col in columnas}
                    fila_sep_lineas['Branch'] = 'LÍNEAS'
                    datos_flujo_post.append(fila_sep_lineas)
                    
                    # Datos de líneas
                    for l in lineas_idx:
                        bus_from = int(branch_from[l])
                        bus_to = int(branch_to[l])
                        fm = float(vars_case['FM'][l])
                        
                        fila = {'Branch': f'{bus_from}-{bus_to}'}
                        
                        for c in range(K):
                            tipo, idx = contingencias[c]
                            flujo = round(float(f_post.X[l, c, w]), 3)
                            cargabilidad = round(abs(flujo) / fm * 100, 2) if fm > 0 else 0
                            
                            if tipo == "gen":
                                base = f"Cont{c+1}: gen. {idx}"
                            else:
                                if idx == 2:
                                    base = f"Cont{c+1}: alm. 2"
                                elif idx == 4:
                                    base = f"Cont{c+1}: alm. 4"
                                elif idx == 11:
                                    base = f"Cont{c+1}: alm. 11"
                                else:
                                    base = f"Cont{c+1}: alm. ENAP"
                            
                            fila[f"{base} - Flujo (MW)"] = flujo
                            fila[f"{base} - Carg (%)"] = cargabilidad
                        
                        datos_flujo_post.append(fila)
                
                # TRANSFORMADORES
                if trafos_idx:
                    # Fila separadora con TODAS las columnas vacías
                    fila_sep_trafos = {col: '' for col in columnas}
                    fila_sep_trafos['Branch'] = 'TRANSFORMADORES'
                    datos_flujo_post.append(fila_sep_trafos)
                    
                    # Datos de trafos
                    for l in trafos_idx:
                        bus_from = int(branch_from[l])
                        bus_to = int(branch_to[l])
                        fm = float(vars_case['FM'][l])
                        
                        fila = {'Branch': f'{bus_from}-{bus_to}'}
                        
                        for c in range(K):
                            tipo, idx = contingencias[c]
                            flujo = round(float(f_post.X[l, c, w]), 3)
                            cargabilidad = round(abs(flujo) / fm * 100, 2) if fm > 0 else 0
                            
                            if tipo == "gen":
                                base = f"Cont{c+1}: gen. {idx}"
                            else:
                                if idx == 2:
                                    base = f"Cont{c+1}: alm. 2"
                                elif idx == 4:
                                    base = f"Cont{c+1}: alm. 4"
                                elif idx == 11:
                                    base = f"Cont{c+1}: alm. 11"
                                else:
                                    base = f"Cont{c+1}: alm. ENAP"
                            
                            fila[f"{base} - Flujo (MW)"] = flujo
                            fila[f"{base} - Carg (%)"] = cargabilidad
                        
                        datos_flujo_post.append(fila)
                
                # Crear DataFrame con columnas ordenadas
                df_flujo_post_w = pd.DataFrame(datos_flujo_post, columns=columnas)
                df_flujo_post_w.to_excel(writer, sheet_name=nombre_hoja, startrow=fila_actual, startcol=0, index=False)
                
                fila_actual += len(df_flujo_post_w) + 3

# 4. Formateo Final
wb = load_workbook(nombre_archivo)

from openpyxl.styles import Alignment, Font

for hoja in wb.sheetnames:
    ws = wb[hoja]
    
    # Procesar combinación de encabezados PRIMERO
    filas_a_procesar = []
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=1), start=1):
        if row[0].value and "Flujos postcontingencia" in str(row[0].value):
            filas_a_procesar.append(row_idx + 1)  # header está 1 fila después
    
    # Procesar cada tabla de flujos post encontrada
    for header_row_idx in reversed(filas_a_procesar):  # Procesar de abajo hacia arriba
        col_idx = 2  # Columna B (después de "Branch")
        
        # Insertar fila para encabezados principales
        ws.insert_rows(header_row_idx)
        
        # Procesar todas las columnas
        while col_idx <= ws.max_column:
            cell = ws.cell(row=header_row_idx + 1, column=col_idx)
            if cell.value and "- Flujo (MW)" in str(cell.value):
                # Extraer nombre base
                base_name = str(cell.value).replace(" - Flujo (MW)", "")
                
                # Combinar celdas para el encabezado principal
                ws.merge_cells(
                    start_row=header_row_idx, start_column=col_idx,
                    end_row=header_row_idx, end_column=col_idx + 1
                )
                
                # Escribir nombre de contingencia
                merged_cell = ws.cell(row=header_row_idx, column=col_idx)
                merged_cell.value = base_name
                merged_cell.alignment = Alignment(horizontal='center', vertical='center')
                merged_cell.font = Font(bold=True)
                
                # Actualizar subencabezados
                ws.cell(row=header_row_idx + 1, column=col_idx).value = "Flujo (MW)"
                ws.cell(row=header_row_idx + 1, column=col_idx + 1).value = "Carg (%)"
                
                col_idx += 2
            else:
                col_idx += 1
    
    # Ajustar ancho de columnas AL FINAL
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 30)

wb.save(nombre_archivo)
print(f"\nArchivo '{nombre_archivo}' generado exitosamente con ENS, flujos separados y cargabilidad.")

os.startfile(nombre_archivo)
