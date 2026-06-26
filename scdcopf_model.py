import numpy as np
from scipy.stats import truncnorm

cto1 = 50 # 50 # 10.3
cto2 = 19.9 # 19.9
llf23 = 30 # min 23 que provoca R_up
llf13 = 100
llf12 = 100
load_m = 50/50
Carga_m = 50 * load_m
k_res = 1#50*0.6 #0.6 #1.5 off gen. 1
k_res2 = 1#19.9*1.5 #1.5

p_fore = 50
alfa = 0.1

# Opt. pre + post (True), False: opt. solo pre
opt_post = True

# Traslado de dda a bus 1
dda_bus1 = False

exp_val = False

if not exp_val:
      #########################################################
      usar_semilla = True

      rng = np.random.default_rng(42 if usar_semilla else None)

      valores = np.round(rng.uniform(-1, 1, 5), 2)
      #print(valores, sum(valores))

      #########################################################
      rng = np.random.default_rng(42)

      # Normal truncada en [-3, 3]
      a, b = -3, 3

      u = truncnorm.rvs(a, b, loc=0, scale=1, size=10, random_state=rng)

      # Normalizar a [-1,1]
      numeros = u / 3 # u_norm

      print(len(numeros))

      #print(np.round(numeros, 2))
      print("Media:", round(np.mean(numeros),5), "Std (poblacional):", 
            round(np.std(numeros),5), "Std (muestral):", round(np.std(numeros, ddof=1),5))

      indice_fore = np.argmin(np.abs(numeros))
      valor_fore = numeros[indice_fore]
      print("Valor cercano al esperado", indice_fore, valor_fore)
      idx_min = np.argmin(np.abs(numeros + 1))
      val_min = numeros[idx_min]
      print("Valor mínimo", idx_min, val_min)
      idx_max = np.argmin(np.abs(numeros - 1))
      val_max = numeros[idx_max]
      print("Valor máximo", idx_max, val_max)

from IPython import get_ipython

# Limpiar la consola
ipython = get_ipython()
ipython.run_line_magic('clear', '')  # limpia la consola
from IPython import get_ipython

import warnings

warnings.filterwarnings("ignore", category = DeprecationWarning)

from gurobipy import *
#from scipy.sparse import csr_matrix as sparse
import numpy as np
from numpy import pi#, array, ones, zeros, arange, ix_, r_, flatnonzero as finddiag, dot as mult
#from numpy.linalg import solve, inv
#import pandas as pd
import time
#import matplotlib.pyplot as plt
#import matplotlib.gridspec as gridspec

import case3b_paper as mpc
sep = mpc.case3b()

t0=time.time()

Sb = sep['baseMVA']
SL = sep['SL'][0]          # slack bus

ng = len(sep['gen'])       # número de gen.
nb = len(sep['bus'])       # número de barras
nl = len(sep['branch'])    # número de líneas

#####################################################################################################################################
# Modificar Pmax gen wind (valor base: 50 MW)
#sep['gen'][1,8] = 100
# Modificar Pmax gen. 3
#sep['gen'][2,8] = 80

# Modificar vector de demanda
if dda_bus1:
    sep['bus'][2,2] = 0
    sep['bus'][0,2] = 50 

Pmax = sep['gen'][:,8]
Pmin = sep['gen'][:,9]

# Ctos. generadores
a_g = sep['gencost'][:,5]
b_g = sep['gencost'][:,6]

# Modificar ctos en gen 1 y gen 3
sep['gencost'][0,5] = cto1
sep['gencost'][2,5] = cto2

# Límite reservas
RUp = RDn = Pmax - Pmin    # límites mín./máx.

# Cto. reservas
Cto_up = np.array([1*k_res, 0.15, 0.7*k_res2])
Cto_dn = np.array([0.25, 0.15, 0.1])

Cg = sep['Cg']             # matriz de conexiones

#####################################################################################################################################
# Modificar FM línea 2 - 3 (valor base: 30 MW)
sep['branch'][2,5] = llf23
# Modificar FM línea 1 - 3 (valor base: 100 MW)
sep['branch'][1,5] = llf13
# Modificar FM línea 1 - 2 (valor base: 100 MW)
sep['branch'][0,5] = llf12

FM = sep['branch'][:,5]    # F^M
A = sep['S']               # matriz de incidencia
Bbus = sep['Bbus']         # B bus
Bf = sep['Bf']             # yA

# Carga pre contingencia
Load_bus_pre = load_m * sep['bus'][:,2]

# ENS
Cto_ens = 500
Pmax_ens = Load_bus_pre

# Parámetros salida de líneas
sl_matrix = np.ones((nl,nl))-np.eye(nl)

# Parámetro binario de generadores que participan en AGC
vf = np.array([1, 1, 1])

contingencias = [('gen',1), ('gen',2), ('gen',3), ('line', 1), ('line', 2), ('line', 3)]
K = len(contingencias)

#####################################################################################################################################
#⚡⚡⚡⚡⚡⚡⚡⚡⚡⚡⚡⚡⚡⚡⚡⚡⚡⚡⚡⚡⚡⚡⚡⚡⚡⚡⚡⚡
# Parámetros escenarios de epsilon
sigma = alfa * p_fore  # desviación estándar
zeta = 3 * sigma       # límite de desviación
p_VUL = p_fore - zeta  # límite de la parte despachable

if exp_val:
    epsilon_list = np.array([0]) * zeta
else:
    epsilon_list = numeros * zeta
 
n_w = len(epsilon_list)

eta_list = zeta + epsilon_list # parte estocástica

e = 1  # generador renovable
eta_vector = np.zeros(ng)

RUp[e] = RDn[e] = p_VUL - Pmin[e]

#####################################################################################################################################
#####################################################################################################################################

# ==== Modelo ====
m=Model('DCOPF_3b')
m.setParam('OutputFlag', False)

# Definición de variables
# Variables de primera etapa - reserva
r_up=m.addMVar(ng, vtype=GRB.CONTINUOUS, lb=0, name='r_up')
r_dn=m.addMVar(ng, vtype=GRB.CONTINUOUS, lb=0, name='r_dn')

# Variable de encendido de gen.
u_i = m.addMVar(ng, vtype=GRB.BINARY, name='u_i')
m.addConstr(u_i[e] == 1, name='u_i_ones')  

# Variables de segunda etapa
# Precontingencia
p_pre=m.addMVar((ng, n_w), vtype=GRB.CONTINUOUS, lb=0, name='Pg_pre') #lb lim inf p>=0
d_pre=m.addMVar((nb, n_w), vtype=GRB.CONTINUOUS, lb=-GRB.INFINITY, ub=GRB.INFINITY, name='d_pre')
f_pre=m.addMVar((nl, n_w), vtype=GRB.CONTINUOUS, lb=-GRB.INFINITY, ub=GRB.INFINITY, name='f_pre')
p_ens_pre=m.addMVar((nb, n_w), vtype=GRB.CONTINUOUS, lb=0, name='p_ens_pre')

# Postcontingencia
p_post=m.addMVar((ng, K, n_w), vtype=GRB.CONTINUOUS, lb=0, name='Pg_post') 
d_post=m.addMVar((nb, K, n_w), vtype=GRB.CONTINUOUS, lb=-GRB.INFINITY, ub=GRB.INFINITY, name='d_post')
f_post=m.addMVar((nl, K, n_w), vtype=GRB.CONTINUOUS, lb=-GRB.INFINITY, ub=GRB.INFINITY, name='f_post')
p_ens_post=m.addMVar((nb, K, n_w), vtype=GRB.CONTINUOUS, lb=0, name='p_ens_post')

#####################################################################################################################################
# ==== Función objetivo ====

Cop = LinExpr()
prob_w = 1 / n_w  # equiprobables

# Cto. pre-contingencia
Cop_pre_total = quicksum(
    a_g @ p_pre[:, w] + b_g @ u_i + Cto_ens * p_ens_pre[:, w].sum()
    for w in range(n_w)
)
# Cto. post-contingencia
if opt_post:
    Cop_post_total = quicksum(
        (a_g @ p_post[:, k, w] + b_g @ u_i + Cto_ens * p_ens_post[:, k, w].sum()) / K
        for w in range(n_w)
        for k in range(K)
    )
else:
    Cop_post_total = quicksum(
        (Cto_ens * p_ens_post[:, k, w].sum()) / K
        for w in range(n_w)
        for k in range(K)
    )

# Contribución del caso w a la FO total
Cop += prob_w * (Cop_pre_total + Cop_post_total)

# ==== Subject to ====

# Reserva
m.addConstr(-r_up >= -RUp * vf, name = 'RUp')
m.addConstr(-r_dn >= -RDn * vf, name = 'RDn')

for w in range(n_w):
    eta = eta_list[w]
    eta_vector[e] = eta

    p_pre_w = p_pre[:, w]
    d_pre_w = d_pre[:, w]
    p_ens_pre_w = p_ens_pre[:, w]
    f_pre_w = f_pre[:, w]

    #####################################################################################################################################
    # ==== Precontingencia ====

    # Barra SL
    m.addConstr(d_pre[SL] == 0, f'SL_pre_w{w}')

    # Balance (LCK)
    m.addConstr(Cg @ p_pre_w + Cg @ eta_vector + p_ens_pre_w - Load_bus_pre == A.T @ f_pre_w, name = f'LCK_pre_w{w}') #Cg*Pg+P_ens-D=Bbus*d

    # Límite líneas
    m.addConstr(f_pre_w == Sb * Bf @ d_pre_w, name = f'f_pre_w{w}')
    m.addConstr(-f_pre_w >= -FM, name = f'fp_pre_w{w}')
    m.addConstr(f_pre_w >= -FM, name = f'fn_pre_w{w}')

    # Límite ángulos
    m.addConstr(-A @ d_pre_w >= -pi/2, name = f'dM_pre_w{w}')
    m.addConstr(A @ d_pre_w >= -pi/2, name = f'dm_pre_w{w}')

    # P_max y P_min
    for t in range(ng):
        if t != e:
            m.addConstr(p_pre_w[t] >= Pmin[t] * u_i[t], name = f'P_min_pre{t}_w{w}')
            m.addConstr(-p_pre_w[t] >= -Pmax[t] * u_i[t], name= f'P_max_pre{t}_w{w}')
    
    m.addConstr(p_pre_w[e] + eta_vector[e] >= Pmin[e], name = f'P_min_pre{e}_w{w}')
    m.addConstr(-p_pre_w[e] - eta_vector[e] >= -Pmax[e], name= f'P_max_pre{e}_w{w}')

    # Límite P_VUL para generador renovable
    m.addConstr(-p_pre_w[e] - r_up[e] >= -p_VUL, name=f'P_VUL_e{e}_w{w}')

    # ENS
    m.addConstr(-p_ens_pre_w >= -Pmax_ens, name= f'P_max_ens_pre_w{w}')

    #####################################################################################################################################
    # ==== Postcontingencia ====

    for k_idx,(tipo,index) in enumerate(contingencias):
        p_post_k     = p_post[:, k_idx, w]
        d_post_k     = d_post[:, k_idx, w]
        p_ens_post_k = p_ens_post[:, k_idx, w]
        f_post_k = f_post[:, k_idx, w]
        eta_vector_post = eta_vector.copy()

        # N-1 cargas
        if tipo == 'load':
            vc_matrix = np.ones((nb,nb))-np.eye(nb)
            vc = vc_matrix[:,index-1]
            Load_bus_post_k = Load_bus_pre.copy() * vc # carga post
        
        # N-1 gen.
        elif tipo == 'gen':
            # Anular el componente estocástico para el generador fuera de servicio
            Load_bus_post_k = Load_bus_pre.copy()
            eta_vector_post[index-1] = 0
        
        # N-1 líneas
        else:
            Load_bus_post_k = Load_bus_pre.copy()

        # Barra SL
        m.addConstr(d_post_k[SL] == 0, f'SL_post[{k_idx}]_w{w}')

        # Balance (LCK)
        m.addConstr(Cg @ p_post_k + Cg @ eta_vector_post + p_ens_post_k - Load_bus_post_k == A.T @ f_post_k, name = f'LCK_post[{k_idx}]_w{w}')

        # Límite líneas
        if tipo == 'line':
            sl = sl_matrix[:,index-1]
            m.addConstr(f_post_k == sl * Sb * (Bf @ d_post_k), name = f'f_post[{k_idx}]_w{w}')
            m.addConstr(-f_post_k >= -FM, name = f'fp_post[{k_idx}]_w{w}')
            m.addConstr(f_post_k >= -FM, name = f'fn_post[{k_idx}]_w{w}')
            
        else:
            m.addConstr(f_post_k == Sb * Bf @ d_post_k, name = f'f_post[{k_idx}]_w{w}')
            m.addConstr(-f_post_k >= -FM, name = f'fp_post[{k_idx}]_w{w}')
            m.addConstr(f_post_k >= -FM, name = f'fn_post[{k_idx}]_w{w}')
            
        # Límite ángulos
        m.addConstr(-A @ d_post_k >= -pi/2, name = f'dM_post[{k_idx}]_w{w}')
        m.addConstr(A @ d_post_k >= -pi/2, name = f'dm_post[{k_idx}]_w{w}')

        # Generador fuera de servicio
        if tipo == 'gen':
            m.addConstr(p_post_k[index-1] == 0, f'Out_service[{k_idx}]_w{w}')
            for h in range(ng):
                if h != index-1:   # todos excepto el fuera de servicio
                    # P_max y P_min
                    if h != e:
                        m.addConstr(p_post_k[h] >= Pmin[h] * u_i[h], name=f'Pmin_post[{k_idx},{h}]_w{w}')
                        m.addConstr(-p_post_k[h] >= -Pmax[h] * u_i[h], name=f'Pmax_post[{k_idx},{h}]_w{w}')
                    if h == e:
                        m.addConstr(p_post_k[h] + eta_vector_post[h] >= Pmin[h], name=f'Pmin_post[{k_idx},{h}]_w{w}')
                        m.addConstr(-p_post_k[h] - eta_vector_post[h] >= -Pmax[h], name=f'Pmax_post[{k_idx},{h}]_w{w}')
                        
                    # Limite p_VUL para generador renovable
                        m.addConstr(-p_post_k[h] >= -p_VUL, name=f'p_VUL_post[{k_idx},{h}]_w{w}')
                    
                    # Reserva
                    m.addConstr(p_pre_w[h] + r_up[h] >= p_post_k[h], name=f'Up[{k_idx},{h}]_w{w}')
                    m.addConstr(-p_pre_w[h] + r_dn[h] >= -p_post_k[h], name=f'Dn[{k_idx},{h}]_w{w}')
        else:
            # P_max y P_min
            for t in range(ng):
                if t != e:
                    m.addConstr(p_post_k[t] >= Pmin[t] * u_i[t], name = f'P_min_post[{k_idx},{t}]_w{w}')
                    m.addConstr(-p_post_k[t] >= -Pmax[t] * u_i[t], name= f'P_max_post[{k_idx},{t}]_w{w}')
            
            m.addConstr(p_post_k[e] + eta_vector_post[e] >= Pmin[e], name = f'P_min_post[{k_idx},{e}]_w{w}')
            m.addConstr(-p_post_k[e] - eta_vector_post[e] >= -Pmax[e], name= f'P_max_post[{k_idx},{e}]_w{w}')
            
            # Limite p_VUL para generador renovable
            m.addConstr(-p_post_k[e] >= -p_VUL, name= f'p_vul_post[{k_idx}]_w{w}')

            # Reserva
            m.addConstr(p_pre_w + r_up >= p_post_k, name = f'Up[{k_idx}]_w{w}')
            m.addConstr(-p_pre_w + r_dn >= -p_post_k, name = f'Dn[{k_idx}]_w{w}')

        # ENS
        m.addConstr(-p_ens_post_k >= -Pmax_ens, name=f'P_max_ens_post[{k_idx}]_w{w}')


C_res = r_up @ Cto_up + r_dn @ Cto_dn
    
Cop += C_res
m.setObjective(Cop, GRB.MINIMIZE)

t1=time.time()
m.optimize()
t2=time.time()

status = m.Status
if status == GRB.Status.OPTIMAL:
    print('Optimal found => status "%d"' % status)
elif status == GRB.Status.INF_OR_UNBD or \
    status == GRB.Status.INFEASIBLE  or \
    status == GRB.Status.UNBOUNDED:
    print('The model cannot be solved because it is infeasible or unbounded => status "%d"' % status)

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd
import os

# ==== Exportar a Excel ====
nombre_archivo = "resultados_paper_escenarios.xlsx"

# Calcular componentes del costo total
Cop_pre_val = sum(
    (a_g @ p_pre[:, w].X + b_g @ u_i.X + Cto_ens * p_ens_pre[:, w].X.sum()) / n_w
    for w in range(n_w)
)

Cop_post_val = sum(
    (a_g @ p_post[:, k, w].X + b_g @ u_i.X + Cto_ens * p_ens_post[:, k, w].X.sum()) / (n_w * K)
    for w in range(n_w)
    for k in range(K)
)

C_res_val = sum(Cto_up[g] * r_up.X[g] + Cto_dn[g] * r_dn.X[g] for g in range(ng))

# ==== Crear una hoja por cada escenario ====
with pd.ExcelWriter(nombre_archivo, engine="openpyxl") as writer:
    
    # Iterar sobre TODOS los escenarios
    for w_idx in range(n_w):
        eta_val = eta_list[w_idx]
        nombre_hoja = f"eta_{eta_val:.3f}"
        
        # ----------------------------------------
        # 1. Tabla precontingencia (generación)
        # ----------------------------------------
        datos_pre = []
        for h in range(ng):
            datos_pre.append({
                "Generador": h+1,
                "Potencia pre (MW)": round(p_pre.X[h, w_idx], 3),
                "Reserva Up (MW)": round(r_up.X[h], 3),
                "Reserva Dn (MW)": round(r_dn.X[h], 3)
            })
        tabla_pre = pd.DataFrame(datos_pre)
        
        # ----------------------------------------
        # 2. Tabla postcontingencia (generación)
        # ----------------------------------------
        datos_post = {"Generador\ contingencia": [h+1 for h in range(ng)]}
        for c in range(K):
            tipo, idx = contingencias[c]
            if tipo == "gen":
                nombre_col = f"Cont{c+1}: gen. {idx}"
            elif tipo == "line":
                nombre_col = f"Cont{c+1}: línea {int(sep['branch'][idx-1,0])}-{int(sep['branch'][idx-1,1])}"
            else:
                nombre_col = f"Cont{c+1}: {tipo}{idx}"
            datos_post[nombre_col] = [round(p_post.X[h, c, w_idx], 3) for h in range(ng)]
        tabla_post = pd.DataFrame(datos_post)

        # ----------------------------------------
        # 3. Tabla ENS (corte de carga)
        # ----------------------------------------
        datos_ens = []
        umbral_cero = 1e-5
        for b in range(nb):
            # Extraer valores para verificar si hay ENS en esta barra
            valor_pre = round(float(p_ens_pre.X[b, w_idx]), 3)
            valores_post = [round(float(p_ens_post[b, c, w_idx].X), 3) for c in range(K)]

            # Condición: ¿Existe algún valor mayor a cero en pre o en cualquier contingencia?
            if valor_pre > umbral_cero or any(v > umbral_cero for v in valores_post):
                fila_ens = {
                    "Bus": b+1, 
                    "ENS Pre": valor_pre
                }
                
                for c in range(K):
                    tipo, idx = contingencias[c]
                    if tipo == "gen":
                        nombre_col = f"Cont{c+1}: gen. {idx}"
                    elif tipo == "line":
                        nombre_col = f"Cont{c+1}: línea {int(sep['branch'][idx-1,0])}-{int(sep['branch'][idx-1,1])}"
                    else:
                        nombre_col = f"Cont{c+1}: {tipo}{idx}"
                    fila_ens[nombre_col] = valores_post[c]
            
                datos_ens.append(fila_ens)
            
        tabla_ens = pd.DataFrame(datos_ens)
        
        # ----------------------------------------
        # 4. Tabla flujos precontingencia
        # ----------------------------------------
        datos_flujos_pre = {
            "Línea": [f"{int(sep['branch'][l,0])}-{int(sep['branch'][l,1])}" for l in range(nl)],
            "Flujo pre (MW)": [round(f_pre.X[l, w_idx], 3) for l in range(nl)]
        }
        tabla_flujos_pre = pd.DataFrame(datos_flujos_pre)

        # ----------------------------------------
        # 5. Tabla flujos POST-contingencia
        # ----------------------------------------
        datos_flujos_post = {
            "Línea \ contingencia": [f"{int(sep['branch'][l,0])}-{int(sep['branch'][l,1])}" for l in range(nl)]
        }
        for c in range(K):
            tipo, idx = contingencias[c]
            if tipo == "gen":
                nombre_col = f"Cont{c+1}: gen. {idx}"
            elif tipo == "line":
                nombre_col = f"Cont{c+1}: línea {int(sep['branch'][idx-1,0])}-{int(sep['branch'][idx-1,1])}"
            else:
                nombre_col = f"Cont{c+1}: {tipo}{idx}"
            
            datos_flujos_post[nombre_col] = [round(f_post.X[l, c, w_idx], 3) for l in range(nl)]
            
        tabla_flujos_post = pd.DataFrame(datos_flujos_post)
        
        # ----------------------------------------
        # Escribir las tablas en la hoja
        # ----------------------------------------
        fila_actual = 0
        
        # 1. Generación pre
        tabla_pre.to_excel(writer, sheet_name=nombre_hoja, index=False, startrow=fila_actual)
        fila_actual += len(tabla_pre) + 2
        
        # 2. Generación post
        tabla_post.to_excel(writer, sheet_name=nombre_hoja, index=False, startrow=fila_actual)
        fila_actual += len(tabla_post) + 2
        
        # 3. ENS
        if not tabla_ens.empty:
            tabla_ens.to_excel(writer, sheet_name=nombre_hoja, index=False, startrow=fila_actual)
            fila_actual += len(tabla_ens) + 2

        # 4. Flujos pre
        tabla_flujos_pre.to_excel(writer, sheet_name=nombre_hoja, index=False, startrow=fila_actual)
        fila_actual += len(tabla_flujos_pre) + 2

        # 5. Flujos post
        tabla_flujos_post.to_excel(writer, sheet_name=nombre_hoja, index=False, startrow=fila_actual)

# ==== Agregar resumen de costos y formato ====
wb = load_workbook(nombre_archivo)

for w_idx in range(n_w):
    eta_val = eta_list[w_idx]
    nombre_hoja = f"eta_{eta_val:.3f}"
    ws = wb[nombre_hoja]
    
    # Calcular fila resumen automáticamente buscando la última fila con datos
    fila_resumen = ws.max_row + 2
    
    # Agregar información del escenario
    ws.cell(row=fila_resumen, column=1, value=f"Escenario: eta = {eta_val:.3f} MW, error = {epsilon_list[w_idx]:.3f} MW")
    ws.cell(row=fila_resumen + 1, column=1, value=f"P_fore = {p_fore:.3f} MW, P_VUL = {p_VUL:.3f} MW")
    
    fila_resumen += 3
    
    # Agregar resumen de costos
    ws.cell(row=fila_resumen, column=1, value="Resumen de costos ($/h)")
    ws.cell(row=fila_resumen + 1, column=1, value="Costo total del sistema")
    ws.cell(row=fila_resumen + 1, column=2, value=round(m.objVal, 3))
    
    ws.cell(row=fila_resumen + 2, column=1, value=f"Costo precontingencia (valor esperado)")
    ws.cell(row=fila_resumen + 2, column=2, value=round(Cop_pre_val, 3))
    
    ws.cell(row=fila_resumen + 3, column=1, value="Costo de reserva (total)")
    ws.cell(row=fila_resumen + 3, column=2, value=round(C_res_val, 3))

    ws.cell(row=fila_resumen + 4, column=1, value="Costo de reserva Up")
    ws.cell(row=fila_resumen + 4, column=2, value=round(Cto_up @ r_up.X, 3))

    ws.cell(row=fila_resumen + 5, column=1, value="Costo de reserva Dn")
    ws.cell(row=fila_resumen + 5, column=2, value=round(Cto_dn @ r_dn.X, 3))
    
    ws.cell(row=fila_resumen + 6, column=1, value=f"Costo postcontingencia (valor esperado)")
    ws.cell(row=fila_resumen + 6, column=2, value=round(Cop_post_val, 3))
    
    # --- Ajustar ancho de columnas ---
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    val_len = len(str(cell.value))
                    if val_len > max_length: max_length = val_len
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[col_letter].width = adjusted_width

wb.save(nombre_archivo)
print(f"\nResultados exportados a: {nombre_archivo}")
print(f"Hojas creadas: {[sheet for sheet in wb.sheetnames]}")

# ==== Abrir el archivo automáticamente ====
#os.startfile(nombre_archivo)
#print(f"Abriendo {nombre_archivo}...")

# Gráficos de barras (sistema de prueba, case3b): potencia y flujos,
# pre y post-contingencia — con exportación a PDF
# ----------------------------------------------------------------------------

import os
import matplotlib.pyplot as plt
import numpy as np

# ============================================================================
# CONFIGURACIÓN DEL USUARIO
# ============================================================================

FIGSIZE = (8, 5)          # ancho, alto [pulgadas] -> tamaño físico real del PDF
DPI = 300                 # resolución para elementos rasterizados (poco relevante en PDF, vectorial)
CARPETA_SALIDA = "graficos_sistema_prueba"
MOSTRAR_EN_PANTALLA = True  # además de exportar, abre las figuras en pantalla

ordenar_escenarios = False  # True: ordena por inyección renovable total

os.makedirs(CARPETA_SALIDA, exist_ok=True)

if status == GRB.Status.OPTIMAL:

    # ---- Extracción de valores óptimos ----
    p_pre_val  = p_pre.X    # (ng, n_w)
    f_pre_val  = f_pre.X    # (nl, n_w)
    p_post_val = p_post.X   # (ng, K, n_w)
    f_post_val = f_post.X   # (nl, K, n_w)

    k_cont = 0  # primera contingencia de la lista `contingencias`

    # ---- Etiquetas de elementos ----
    gen_labels_base = [f'Gen {i+1}' for i in range(ng)]
    line_labels = [f"Línea {int(sep['branch'][l, 0])}-{int(sep['branch'][l, 1])}" for l in range(nl)]

    # ---- INCORPORACIÓN DE LA PARTE ESTOCÁSTICA (eta) ----
    e_idx = 1  # Índice del generador renovable (Gen 2)

    p_pre_real = p_pre_val.copy()
    p_post_real = p_post_val.copy()

    for w in range(n_w):
        p_pre_real[e_idx, w] += eta_list[w]
        for k_idx, (tipo, index) in enumerate(contingencias):
            if tipo == 'gen' and (index - 1) == e_idx:
                p_post_real[e_idx, k_idx, w] = 0
            else:
                p_post_real[e_idx, k_idx, w] += eta_list[w]

    # ---- CONDICIONAL DE ORDENAMIENTO ----
    if ordenar_escenarios:
        sort_idx = np.argsort(p_pre_real[e_idx, :])
    else:
        sort_idx = np.arange(n_w)

    x_pos = np.arange(n_w)
    etiquetas_x = [f'{w+1}' for w in sort_idx]

    def etiquetas_con_contingencia(k_idx, base_gen, base_line):
        tipo, idx = contingencias[k_idx]
        gl, ll = base_gen.copy(), base_line.copy()
        if tipo == 'gen':
            gl[idx - 1] += ' (f.s.)'
        elif tipo == 'line':
            ll[idx - 1] += ' (f.s.)'
        return gl, ll

    # ========================================================================
    # PALETA DE COLORES (ampliada)
    # ----------------------------------------------------------------------
    # - Térmicas / otras unidades no-ERV: marrones, naranjos, rojos (sin cambios)
    # - Líneas: azules y morados (ampliado; antes incluía un tono cian que no
    #   correspondía a la familia pedida, se reemplazó)
    # - WTG eólicos: verdes (ampliado a 3 variantes, una por unidad eólica)
    # ========================================================================
    colores_termicos = ['#7C4B3A', '#7D6642', '#B8643B', '#B78D43', '#dd8452', '#C44E52']
    colores_wtg       = ['#55a868', '#3C8F5C', '#8FCB9B']   # 3 variantes de verde
    colores_fv         = ['#F2CC77']                        # dorado, fotovoltaico (si aplica)
    colores_lineas_pool = [
        '#4C72B0', '#8172B3', '#6A8FC4', '#A084CA',
        '#2E5090', '#5E3C99', '#7B9CD4', '#9B6FB5',
    ]  # azules y morados alternados; agregar más entradas aquí si se necesitan

    # Asignación puntual para este sistema de 3 gen. / 3 líneas
    colores_gen = [colores_termicos[0], colores_wtg[0], colores_termicos[4]]
    colores_lineas = colores_lineas_pool[:3]

    figs_a_exportar = []  # (fig, nombre_archivo)

    # ========================================================================
    # 1. POTENCIA - PRE-CONTINGENCIA (Barras apiladas)
    # ========================================================================
    fig1, ax1 = plt.subplots(figsize=FIGSIZE)
    bottom = np.zeros(n_w)

    for i in range(ng):
        y_data = p_pre_real[i, sort_idx]
        ax1.bar(x_pos, y_data, bottom=bottom, label=gen_labels_base[i],
                color=colores_gen[i], edgecolor='black', linewidth=0.7, width=0.6)
        bottom += y_data

    ax1.axhline(y=50, color='#4C72B0', linestyle='--', linewidth=1, label='Demanda (50 MW)')
    ax1.set_xticks(x_pos)
    ax1.set_xticklabels(etiquetas_x)
    ax1.set_xlabel('Escenarios de incertidumbre (w)')
    ax1.set_ylabel('Potencia [MW]')
    ax1.set_title('Despachos pre-contingencia')
    ax1.legend(loc='upper center', bbox_to_anchor=(0.5, -0.15), ncol=ng + 1)
    ax1.grid(True, axis='y', alpha=0.3)
    plt.tight_layout()
    figs_a_exportar.append((fig1, 'pre_potencia.pdf'))

    # ========================================================================
    # 2. FLUJOS - PRE-CONTINGENCIA (Barras agrupadas)
    # ========================================================================
    fig2, ax2 = plt.subplots(figsize=FIGSIZE)
    ancho_barra = 0.8 / nl

    for l in range(nl):
        offset = (l - nl/2 + 0.5) * ancho_barra
        ax2.bar(x_pos + offset, f_pre_val[l, sort_idx], width=ancho_barra,
                label=line_labels[l], color=colores_lineas[l], edgecolor='black', linewidth=0.7)

    ax2.axhline(0, color='black', linewidth=1)
    ax2.set_xticks(x_pos)
    ax2.set_xticklabels(etiquetas_x)
    ax2.set_xlabel('Escenarios de incertidumbre (w)')
    ax2.set_ylabel('Flujos [MW]')
    ax2.set_title('Flujos pre-contingencia')
    ax2.legend(loc='upper center', bbox_to_anchor=(0.5, -0.15), ncol=nl)
    ax2.grid(True, axis='y', alpha=0.3)
    plt.tight_layout()
    figs_a_exportar.append((fig2, 'pre_flujos.pdf'))

    # ========================================================================
    # 3. POTENCIA - POST-CONTINGENCIA (k_cont)
    # ========================================================================
    fig3, ax3 = plt.subplots(figsize=FIGSIZE)
    gen_labels_k, line_labels_k = etiquetas_con_contingencia(k_cont, gen_labels_base, line_labels)
    tipo_k, idx_k = contingencias[k_cont]

    bottom = np.zeros(n_w)
    for i in range(ng):
        y_data = p_post_real[i, k_cont, sort_idx]
        ax3.bar(x_pos, y_data, bottom=bottom, label=gen_labels_k[i],
                color=colores_gen[i], edgecolor='black', linewidth=0.7, width=0.6)
        bottom += y_data

    ax3.axhline(y=50, color='#4C72B0', linestyle='--', linewidth=1, label='Demanda (50 MW)')
    ax3.set_xticks(x_pos)
    ax3.set_xticklabels(etiquetas_x)
    ax3.set_xlabel('Escenarios de incertidumbre (w)')
    ax3.set_ylabel('Potencia [MW]')
    ax3.set_title(f'Despachos post-contingencia (salida {tipo_k} {idx_k})')
    ax3.legend(loc='upper center', bbox_to_anchor=(0.5, -0.15), ncol=ng + 1)
    ax3.grid(True, axis='y', alpha=0.3)
    plt.tight_layout()
    figs_a_exportar.append((fig3, f'post_{tipo_k}{idx_k}_potencia.pdf'))

    # ========================================================================
    # 4. FLUJOS - POST-CONTINGENCIA (k_cont)
    # ========================================================================
    fig4, ax4 = plt.subplots(figsize=FIGSIZE)

    for l in range(nl):
        offset = (l - nl/2 + 0.5) * ancho_barra
        ax4.bar(x_pos + offset, f_post_val[l, k_cont, sort_idx], width=ancho_barra,
                label=line_labels_k[l], color=colores_lineas[l], edgecolor='black', linewidth=0.7)

    ax4.axhline(0, color='black', linewidth=1)
    ax4.set_xticks(x_pos)
    ax4.set_xticklabels(etiquetas_x)
    ax4.set_xlabel('Escenarios de incertidumbre (w)')
    ax4.set_ylabel('Flujos [MW]')
    ax4.set_title(f'Flujos post-contingencia (salida {tipo_k} {idx_k})')
    ax4.legend(loc='upper center', bbox_to_anchor=(0.5, -0.15), ncol=nl)
    ax4.grid(True, axis='y', alpha=0.3)
    plt.tight_layout()
    figs_a_exportar.append((fig4, f'post_{tipo_k}{idx_k}_flujos.pdf'))

    # ========================================================================
    # EXPORTACIÓN A PDF
    # ========================================================================
    for fig, nombre in figs_a_exportar:
        ruta = os.path.join(CARPETA_SALIDA, nombre)
        fig.savefig(ruta, format='pdf', dpi=DPI, bbox_inches='tight')
        print(f'Guardado: {ruta}')

    if MOSTRAR_EN_PANTALLA:
        plt.show()
    else:
        for fig, _ in figs_a_exportar:
            plt.close(fig)

    print(f'\nTotal: {len(figs_a_exportar)} gráficos exportados en {os.path.abspath(CARPETA_SALIDA)}')

else:
    print('No se generan gráficos: el modelo no alcanzó óptimo (status %d).' % status)
