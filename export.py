#Enlazar python con PowerFactory #llamar al sistema operativo
import os;
os.environ["PATH"]=r"D:\Program Files\DIgSILENT\PowerFactory 2021 SP2"+os.environ["PATH"]

#Importar la aplicación #tener acceso desde la ruta donde esta powerfactory
import sys
sys.path.append(r"D:\Program Files\DIgSILENT\PowerFactory 2021 SP2\Python\3.9")

#Importar la aplicación #Importar el módulo que da acceso a la aplicación
import powerfactory as pf

app=pf.GetApplication()
# app.Show() #Abrir pf en Modo Engine
 
#Activar el proyecto
user =app.GetCurrentUser()                #Abre el usuario
project=app.ActivateProject('BD SM Punta Arenas 2023 ECF y EDAC') #Abre el archivo pfd
prj = app.GetActiveProject()              #Activar el proyecto

Pd_hist = [] # lista para almacenar Pd de cada caso
# Recorrer todos los casos
for i in range(1, 7):   # 1 hasta 6
    if i == 3:
        continue
    nombre_escenario = f"CASO {i}"
    print(f"\n--- {nombre_escenario} ---")
    
    # Seleccionar y activar escenario
    escenario = prj.GetContents(nombre_escenario, 1)[0]
    escenario.Activate()

    #LdF
    lf = app.GetFromStudyCase('ComLdf')
    lf.Execute()
    print('Solución ComLdf:')
    print(' Alimentador 2: P =', round(app.GetCalcRelevantObjects('Alimentador 02.ElmLod')[0].GetAttribute('plini'),3), 'MW')
    print(' Alimentador 4: P =', round(app.GetCalcRelevantObjects('Alimentador 04.ElmLod')[0].GetAttribute('plini'),3), 'MW')
    print(' Alimentador 11: P =', round(app.GetCalcRelevantObjects('Alimentador 11.ElmLod')[0].GetAttribute('plini'),3), 'MW')

    interruptor = app.GetCalcRelevantObjects('Breaker/Switch(12).ElmCoup')[0]
    p1 = interruptor.GetAttribute('m:P:bus1')
    p2 = interruptor.GetAttribute('m:P:bus2')
    print(" Alimentador 6 (Conectado a la barra mediante el interruptor 12):")
    print(f"  Flujo por interruptor (12): f = {round(p1,3)} MW (bus1) / {round(p2,3)} MW (bus2)")

    # dda total
    dda_alm = 0
    loads = app.GetCalcRelevantObjects('*.ElmLod')

    load_bus2 = []
    load_bus3 = []

    for l in loads:
        if l.loc_name.startswith("Alimentador"):
            potencia = l.GetAttribute('plini')
            dda_alm += potencia
            num = int(l.loc_name.split()[1])
            if num in [4, 5, 6, 7, 11, 13]:
                load_bus2.append((l.loc_name, potencia))
            else:
                load_bus3.append((l.loc_name, potencia))
    load_bus2.append(("Alimentador 6", p1))
    print('')
    if load_bus2:
        max_bus2 = max(load_bus2, key= lambda x: x[1])
        print(f"Mayor carga en bus 'Barra principal de 13.2kV': {max_bus2[0]} con {round(max_bus2[1],3)} MW")
    if load_bus3:
        max_bus3 = max(load_bus3, key=lambda x: x[1])
        print(f"Mayor carga en bus 'Celdas G.E. 13.2kV': {max_bus3[0]} con {round(max_bus3[1],3)} MW")

    enap = app.GetCalcRelevantObjects('Planta Cabo Negro ENAP.ElmLod')[0]
    p_enap = enap.GetAttribute('plini')

    ssaa_VP = app.GetCalcRelevantObjects('PE VPatagónicos SSAA.ElmLod')[0]
    p_ssaa_VP = ssaa_VP.GetAttribute('plini')

    Pd = dda_alm + p1 + p_enap + p_ssaa_VP
    print('Planta ENAP: P =', round(p_enap,3), 'MW', round(p_enap * 100/Pd, 3), '%')
    Pd_hist.append(round(Pd, 3))
    print('')
    print('Demanda total del sistema Pd =', round(Pd, 3), 'MW')
    dda_bus2 = sum(valor for nombre, valor in load_bus2)
    dda_bus3 = sum(valor for nombre, valor in load_bus3)
    print(" Demanda en bus 'Barra principal de 13.2kV' =", round(dda_bus2,3), 'MW,', round(dda_bus2 * 100/Pd, 3), '%')
    print(" Demanda en bus 'Celdas G.E. 13.2kV' =", round(dda_bus3,3), 'MW,', round(dda_bus3 * 100/Pd, 3), '%')

print("\nDemanda por escenario")
print(Pd_hist)

-------------------------------------------------------------------------------------------------------------------------

"""
exportar_excel.py
-----------------
Exporta los resultados del flujo de carga (PowerFactory) a un archivo .xlsx.
Requiere que las variables del script principal ya estén definidas:
    load_bus2, load_bus3, p_enap, Pd, dda_bus2, dda_bus3, max_bus2, max_bus3
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ── helpers de estilo ──────────────────────────────────────────────────────────

FONT_NAME = "Arial"
COLOR_HEADER    = "2F5496"   # azul oscuro
COLOR_BUS2      = "D9E1F2"   # azul pálido  → Barra principal 13.2 kV
COLOR_BUS3      = "E2EFDA"   # verde pálido → Celdas G.E. 13.2 kV
COLOR_ENAP      = "FFF2CC"   # amarillo pálido → ENAP (sin barra)
COLOR_SUBTOTAL  = "F2F2F2"   # gris claro
COLOR_NOTA      = "FFFFFF"

thin = Side(style="thin", color="AAAAAA")
medium = Side(style="medium", color="555555")
border_thin   = Border(left=thin,   right=thin,   top=thin,   bottom=thin)
border_medium = Border(left=medium, right=medium, top=medium, bottom=medium)
border_top_medium = Border(left=thin, right=thin, top=medium, bottom=thin)


def _hdr(ws, row, col, value, bg=COLOR_HEADER, bold=True, center=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name=FONT_NAME, bold=bold, color="FFFFFF" if bg == COLOR_HEADER else "000000")
    c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="center" if center else "left",
                            vertical="center", wrap_text=True)
    c.border = border_thin
    return c


def _cell(ws, row, col, value, bg="FFFFFF", bold=False, center=True, number=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name=FONT_NAME, bold=bold)
    c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="center" if center else "left", vertical="center")
    c.border = border_thin
    if number and value is not None:
        c.number_format = '#,##0.000'
    return c


def _section_label(ws, row, label, bg, ncols=3):
    """Fila de separador con etiqueta de sección (merge de columnas)."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=label)
    c.font = Font(name=FONT_NAME, bold=True, color="000000")
    c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = border_top_medium
    for col in range(2, ncols + 1):
        ws.cell(row=row, column=col).border = border_top_medium


# ── función principal ──────────────────────────────────────────────────────────

def exportar_excel(
    load_bus2, load_bus3, p_enap,
    Pd, dda_bus2, dda_bus3,
    max_bus2, max_bus3,
    nombre_caso="",
    nombre_archivo="resumen_alimentadores.xlsx"
):
    """
    Parámetros
    ----------
    load_bus2   : list of (nombre, potencia)  – alimentadores en Barra principal 13.2 kV
    load_bus3   : list of (nombre, potencia)  – alimentadores en Celdas G.E. 13.2 kV
    p_enap      : float  – potencia Planta ENAP (MW)
    Pd          : float  – demanda total (MW)
    dda_bus2    : float  – demanda bus2 (MW)
    dda_bus3    : float  – demanda bus3 (MW)
    max_bus2    : tuple  (nombre, potencia)  – alimentador de mayor valor en bus2
    max_bus3    : tuple  (nombre, potencia)  – alimentador de mayor valor en bus3
    nombre_caso : str    – nombre del caso (título de la hoja)
    nombre_archivo : str – ruta del .xlsx de salida
    """

    wb = Workbook()
    ws = wb.active
    ws.title = nombre_caso if nombre_caso else "Resultados"
    ws.sheet_view.showGridLines = False

    # ── anchos de columna ──
    ws.column_dimensions["A"].width = 30   # Alimentador
    ws.column_dimensions["B"].width = 18   # P (MW)
    ws.column_dimensions["C"].width = 18   # % del total

    # ══════════════════════════════════════════════════════
    # TABLA 1 – Despacho por alimentador
    # ══════════════════════════════════════════════════════

    # Título
    ws.merge_cells("A1:C1")
    t = ws["A1"]
    t.value = f"Flujo de carga – {nombre_caso}" if nombre_caso else "Flujo de carga"
    t.font = Font(name=FONT_NAME, bold=True, size=12)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Encabezados tabla 1
    row = 2
    _hdr(ws, row, 1, "Alimentador")
    _hdr(ws, row, 2, "P (MW)")
    _hdr(ws, row, 3, "% Demanda total")
    ws.row_dimensions[row].height = 18

    # ── bus2 ──
    row += 1
    _section_label(ws, row,
                   "Barra principal de 13.2 kV",
                   COLOR_BUS2, ncols=3)
    ws.row_dimensions[row].height = 16

    for nombre, pot in load_bus2:
        row += 1
        _cell(ws, row, 1, nombre,           bg=COLOR_BUS2, center=False)
        _cell(ws, row, 2, round(pot, 3),    bg=COLOR_BUS2, number=True)
        _cell(ws, row, 3, f"={get_column_letter(2)}{row}/{Pd:.6f}*100",
              bg=COLOR_BUS2)
        ws.cell(row=row, column=3).number_format = '0.00"%"'
        ws.row_dimensions[row].height = 15

    # ── bus3 ──
    row += 1
    _section_label(ws, row,
                   "Celdas G.E. 13.2 kV",
                   COLOR_BUS3, ncols=3)
    ws.row_dimensions[row].height = 16

    for nombre, pot in load_bus3:
        row += 1
        _cell(ws, row, 1, nombre,           bg=COLOR_BUS3, center=False)
        _cell(ws, row, 2, round(pot, 3),    bg=COLOR_BUS3, number=True)
        _cell(ws, row, 3, f"={get_column_letter(2)}{row}/{Pd:.6f}*100",
              bg=COLOR_BUS3)
        ws.cell(row=row, column=3).number_format = '0.00"%"'
        ws.row_dimensions[row].height = 15

    # ── ENAP (sin barra) ──
    row += 1
    _section_label(ws, row, "Otros", COLOR_ENAP, ncols=3)
    ws.row_dimensions[row].height = 16

    row += 1
    _cell(ws, row, 1, "Planta Cabo Negro ENAP", bg=COLOR_ENAP, center=False)
    _cell(ws, row, 2, round(p_enap, 3),          bg=COLOR_ENAP, number=True)
    _cell(ws, row, 3, f"={get_column_letter(2)}{row}/{Pd:.6f}*100",
          bg=COLOR_ENAP)
    ws.cell(row=row, column=3).number_format = '0.00"%"'
    ws.row_dimensions[row].height = 15

    # ══════════════════════════════════════════════════════
    # TABLA 2 – Resumen de demanda
    # ══════════════════════════════════════════════════════

    row += 2   # separación entre tablas

    # Encabezados tabla 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    h = ws.cell(row=row, column=1, value="Resumen de demanda")
    h.font = Font(name=FONT_NAME, bold=True, color="FFFFFF")
    h.fill = PatternFill("solid", start_color=COLOR_HEADER)
    h.alignment = Alignment(horizontal="center", vertical="center")
    h.border = border_thin
    ws.row_dimensions[row].height = 18

    # Filas de resumen
    resumen = [
        ("Demanda total del sistema",              round(Pd, 3),     ""),
        ("Barra principal de 13.2 kV",             round(dda_bus2, 3),
         f"{dda_bus2 * 100 / Pd:.2f} %"),
        ("Celdas G.E. 13.2 kV",                   round(dda_bus3, 3),
         f"{dda_bus3 * 100 / Pd:.2f} %"),
    ]

    col_bgs = [COLOR_SUBTOTAL, COLOR_BUS2, COLOR_BUS3]

    for i, (label, val, pct) in enumerate(resumen):
        row += 1
        bg = col_bgs[i]
        _cell(ws, row, 1, label, bg=bg, center=False, bold=(i == 0))
        _cell(ws, row, 2, val,   bg=bg, number=True)
        _cell(ws, row, 3, pct,   bg=bg)
        ws.row_dimensions[row].height = 15

    # Nota alimentadores de mayor valor (una sola vez)
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    nota = (
        f"Alimentadores de mayor valor: "
        f"{max_bus2[0]} (Barra principal 13.2 kV) · "
        f"{max_bus3[0]} (Celdas G.E. 13.2 kV)"
    )
    n = ws.cell(row=row, column=1, value=nota)
    n.font = Font(name=FONT_NAME, italic=True, size=9, color="555555")
    n.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    n.border = border_thin
    ws.row_dimensions[row].height = 24

    wb.save(nombre_archivo)
    print(f"Archivo Excel guardado: {nombre_archivo}")


# ══════════════════════════════════════════════════════════════════════════════
# INTEGRACIÓN CON EL SCRIPT PRINCIPAL
# ══════════════════════════════════════════════════════════════════════════════
# Reemplaza el bloque de impresión al final de cada iteración por:
#
#   from exportar_excel import exportar_excel
#
#   exportar_excel(
#       load_bus2    = load_bus2,
#       load_bus3    = load_bus3,
#       p_enap       = p_enap,
#       Pd           = Pd,
#       dda_bus2     = dda_bus2,
#       dda_bus3     = dda_bus3,
#       max_bus2     = max_bus2,
#       max_bus3     = max_bus3,
#       nombre_caso  = nombre_escenario,          # e.g. "CASO 1"
#       nombre_archivo = f"resultados_{nombre_escenario.replace(' ', '_')}.xlsx"
#   )
#
# O bien, si prefieres un único archivo con varias hojas (una por caso),
# instancia el Workbook fuera del loop y adapta la función para recibir
# el workbook como argumento.
# ══════════════════════════════════════════════════════════════════════════════

exportar_excel(
    load_bus2, load_bus3, p_enap,
    Pd, dda_bus2, dda_bus3,
    max_bus2, max_bus3,
    nombre_caso="",
    nombre_archivo="resumen_alimentadores.xlsx"
)

-------------------------------------------------------------------------------------------------------------------------

"""
exportar_latex.py
-----------------
Exporta los resultados del flujo de carga (PowerFactory) como tablas LaTeX.
Genera un archivo .tex con dos tablas por caso:
    1. Despacho por alimentador (con separadores de barra)
    2. Resumen de demanda
"""


def fmt(value: float, decimals: int = 3) -> str:
    """
    Formatea un número para LaTeX:
    - Separador decimal : coma
    - Separador de miles: punto
    - Elimina ceros finales en la parte decimal
    """
    if abs(value) < 10 ** (-decimals - 1):
        return "0"
    formatted = f"{value:,.{decimals}f}".replace(",", "X").replace(".", ",").replace("X", ".")
    if "," in formatted:
        integer_part, decimal_part = formatted.split(",")
        decimal_part = decimal_part.rstrip("0")
        return integer_part if not decimal_part else f"{integer_part},{decimal_part}"
    return formatted


# ── plantilla de tabla LaTeX ───────────────────────────────────────────────────

_TABLE_OPEN = r"""\begin{{table}}[H]
\renewcommand{{\arraystretch}}{{1.3}}
\centering
\caption{{{caption}}}
\label{{tab:{label}}}"""

_TABLE_CLOSE = r"""\end{tabular}
\end{table}"""


def _tabla_alimentadores(load_bus2, load_bus3, p_enap, Pd, caption="", label="alimentadores"):
    """
    Genera la tabla LaTeX de despacho por alimentador.
    Estructura:
        | Alimentador | P (MW) | % Pd |
        ← separador bus2 →
        ... filas bus2 ...
        ← separador bus3 →
        ... filas bus3 ...
        ← separador ENAP →
        ... fila ENAP ...
    """
    lines = []
    lines.append(_TABLE_OPEN.format(caption=caption, label=label))
    lines.append(r"\begin{tabular}{|l|r|r|}")
    lines.append(r"\hline")
    lines.append(r"\textbf{Alimentador} & \textbf{$P$ (MW)} & \textbf{\% $P_d$} \\")
    lines.append(r"\hline")

    # ── bus2 ──
    lines.append(
        r"\multicolumn{3}{|l|}{\cellcolor[HTML]{D9E1F2}"
        r"\textit{Barra principal de 13,2 kV}} \\"
    )
    lines.append(r"\hline")
    for nombre, pot in load_bus2:
        pct = pot * 100 / Pd
        lines.append(
            f"{nombre} & {fmt(pot)} & {fmt(pct, 2)} \\\\"
        )
    lines.append(r"\hline")

    # ── bus3 ──
    lines.append(
        r"\multicolumn{3}{|l|}{\cellcolor[HTML]{E2EFDA}"
        r"\textit{Celdas G.E. 13,2 kV}} \\"
    )
    lines.append(r"\hline")
    for nombre, pot in load_bus3:
        pct = pot * 100 / Pd
        lines.append(
            f"{nombre} & {fmt(pot)} & {fmt(pct, 2)} \\\\"
        )
    lines.append(r"\hline")

    # ── ENAP ──
    lines.append(
        r"\multicolumn{3}{|l|}{\cellcolor[HTML]{FFF2CC}"
        r"\textit{Otros}} \\"
    )
    lines.append(r"\hline")
    pct_enap = p_enap * 100 / Pd
    lines.append(
        f"Planta Cabo Negro ENAP & {fmt(p_enap)} & {fmt(pct_enap, 2)} \\\\"
    )
    lines.append(r"\hline")
    lines.append(_TABLE_CLOSE)
    return "\n".join(lines)


def _tabla_resumen(Pd, dda_bus2, dda_bus3, max_bus2, max_bus3,
                   caption="", label="resumen"):
    """
    Genera la tabla LaTeX de resumen de demanda.
    Incluye nota de los alimentadores de mayor valor (una sola vez).
    """
    lines = []
    lines.append(_TABLE_OPEN.format(caption=caption, label=label))
    lines.append(r"\begin{tabular}{|l|r|r|}")
    lines.append(r"\hline")
    lines.append(r"\textbf{Concepto} & \textbf{$P$ (MW)} & \textbf{\% $P_d$} \\")
    lines.append(r"\hline")

    rows = [
        ("Demanda total del sistema",     Pd,      100.0,         "FFFFFF"),
        ("Barra principal de 13,2 kV",    dda_bus2, dda_bus2 * 100 / Pd, "D9E1F2"),
        ("Celdas G.E. 13,2 kV",           dda_bus3, dda_bus3 * 100 / Pd, "E2EFDA"),
    ]

    for label_row, val, pct, color in rows:
        color_cmd = f"\\cellcolor[HTML]{{{color}}}" if color != "FFFFFF" else ""
        lines.append(
            f"{color_cmd}{label_row} & {fmt(val)} & {fmt(pct, 2)} \\\\"
        )
        lines.append(r"\hline")

    # Nota: alimentadores de mayor valor
    nota = (
        f"Mayor en barra principal: {max_bus2[0]}. "
        f"Mayor en Celdas G.E.: {max_bus3[0]}."
    )
    lines.append(
        r"\multicolumn{3}{|l|}{\small\textit{"
        + nota
        + r"}} \\"
    )
    lines.append(r"\hline")
    lines.append(_TABLE_CLOSE)
    return "\n".join(lines)


# ── función principal ──────────────────────────────────────────────────────────

def exportar_latex(
    load_bus2, load_bus3, p_enap,
    Pd, dda_bus2, dda_bus3,
    max_bus2, max_bus3,
    nombre_caso="",
    nombre_archivo="tablas_latex.tex"
):
    """
    Escribe (o añade) las dos tablas LaTeX de un caso al archivo indicado.

    Parámetros
    ----------
    load_bus2   : list of (nombre, potencia)
    load_bus3   : list of (nombre, potencia)
    p_enap      : float – potencia Planta ENAP (MW)
    Pd          : float – demanda total (MW)
    dda_bus2    : float – demanda bus2 (MW)
    dda_bus3    : float – demanda bus3 (MW)
    max_bus2    : tuple (nombre, potencia)
    max_bus3    : tuple (nombre, potencia)
    nombre_caso : str   – etiqueta del caso (usada en caption/label)
    nombre_archivo : str – ruta del .tex de salida
    """
    caso_id = nombre_caso.lower().replace(" ", "_")

    tabla1 = _tabla_alimentadores(
        load_bus2, load_bus3, p_enap, Pd,
        caption=f"Despacho por alimentador -- {nombre_caso}",
        label=f"alim_{caso_id}"
    )

    tabla2 = _tabla_resumen(
        Pd, dda_bus2, dda_bus3, max_bus2, max_bus3,
        caption=f"Resumen de demanda -- {nombre_caso}",
        label=f"resumen_{caso_id}"
    )

    # Encabezado de sección (solo si se provee nombre de caso)
    seccion = f"\n% ── {nombre_caso} ──────────────────────────────\n" if nombre_caso else ""

    with open(nombre_archivo, "a", encoding="utf-8") as f:
        f.write(seccion)
        f.write(tabla1)
        f.write("\n\n")
        f.write(tabla2)
        f.write("\n\n")

    print(f"Tablas LaTeX de '{nombre_caso}' añadidas a: {nombre_archivo}")


# ══════════════════════════════════════════════════════════════════════════════
# PAQUETES LATEX NECESARIOS (añadir al preámbulo del documento)
# ══════════════════════════════════════════════════════════════════════════════
# \usepackage{float}       % opción [H]
# \usepackage{colortbl}    % \cellcolor
# \usepackage{xcolor}      % colores HTML
# \usepackage{booktabs}    % (opcional, mejora líneas)
# \usepackage{multirow}    % (opcional, celdas multirow)
#
# ── INTEGRACIÓN CON EL SCRIPT PRINCIPAL ──────────────────────────────────────
# Al inicio (fuera del loop):
#
#   import os
#   from exportar_latex import exportar_latex
#   archivo_tex = "tablas_latex.tex"
#   if os.path.exists(archivo_tex):      # limpiar archivo previo
#       os.remove(archivo_tex)
#
# Dentro del loop, al final de cada caso:
#
#   exportar_latex(
#       load_bus2      = load_bus2,
#       load_bus3      = load_bus3,
#       p_enap         = p_enap,
#       Pd             = Pd,
#       dda_bus2       = dda_bus2,
#       dda_bus3       = dda_bus3,
#       max_bus2       = max_bus2,
#       max_bus3       = max_bus3,
#       nombre_caso    = nombre_escenario,
#       nombre_archivo = archivo_tex,
#   )
# ══════════════════════════════════════════════════════════════════════════════
